# AI SALES AGENT
## - Includes:
# - Generate Lead Scope - Google Maps, X, LinkedIn
# - Multi-client SaaS mode
# - Lead scoring & filtering
# - LinkedIn DM generation
# - Excel export + notifications
# - Stripe subscription gating
# - NO-WEBSITE LEAD TARGETING
# - Maps Link Generation
# - Generate Tailored Outreach - AI written SMS copy, and CRM Sync Notion
# - Notion CRM sync (attach prompt, Loom script, pricing)
# - AI-generated Web App Prompt
# - AI-written Loom-style video script - Video consultation
# - AI-generated SMS copy
# - Auto proposal pricing logic

## - Overall Architecture:
# - Google Sheets (lead source + AI prompt storage)
# - Notion CRM sync (attach prompt, Loom script, pricing)
# - AI-generated Web App Prompt
# - AI-written Loom-style video script
# - AI-generated SMS copy
# - Auto proposal pricing logic

# Scrape Lead
# → Detect No Website
# → Generate Tailored Outreach
# → Generate AI Web App Prompt
# → Store in Google Sheet
# → Notify You (Email / SMS / Telegram)


import os
import stripe
import datetime
import openai
import requests
import gspread
import json
import hashlib
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas
from google.oauth2.service_account import Credentials
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from openpyxl import load_workbook
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials, ServiceAccountCredentials
from twilio.rest import Client
from openai import OpenAI
from googleapiclient.discovery import build


# =========================
# CONFIG
# =========================

# ---------------- ENV SETUP ---------------- #
openai.api_key = os.getenv("OPENAI_API_KEY")
openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
SERPAPI_KEY = os.getenv("SERPAPI_API_KEY")
SENDGRID_KEY = os.getenv("SENDGRID_API_KEY")
NOTION_API_KEY = os.getenv("NOTION_API_KEY")
NOTION_DB_ID = os.getenv("NOTION_DATABASE_ID")
SHEET_ID = os.getenv("GOOGLE_SHEET_ID")
SPREADSHEET_NAME = "SALES_AGENT_LEADS"
CALENDAR_ID = "primary"
STATE_FILE = "sheet_state.json" # TRACK ROW COUNT IN GOOGLE SHEET STATE FILE
STATE_FILE = "state.json" # TRACK ROW COUNT IN EXCEL STATE FILE
stripe.api_key = os.getenv("STRIPE_SECRET_KEY")

with open("config.json") as f:
    CONFIG = json.load(f)

# ---------------- GOOGLE AUTH ---------------- #
creds = Credentials.from_service_account_file(
    "service_account.json",
    scopes=[
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/calendar",
    ]
)

gc = gspread.authorize(creds)
sheet = gc.open(SPREADSHEET_NAME).sheet1
calendar = build("calendar", "v3", credentials=creds)

#v2
# scope = [
#     "https://spreadsheets.google.com/feeds",
#     "https://www.googleapis.com/auth/drive"
# ]

# creds = ServiceAccountCredentials.from_json_keyfile_name(
#     "google-service-account.json", scope
# )
# gs_client = gspread.authorize(creds)
# sheet = gs_client.open_by_key(SHEET_ID).sheet1


# =========================
# OPENAI GENERATION
# =========================
def ai_generate(prompt, temperature=0.7):
    response = openai.ChatCompletion.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=temperature
    )
    return response.choices[0].message.content.strip()


# =========================
# AI PROMPT GENERATOR
# =========================
def build_web_app_prompt(lead):
    return f"""
You are a senior conversion-focused product designer and full-stack engineer.

Build a high-converting web app for:
Business: {lead['business']}
Industry: {lead['industry']}
Target Customer: {lead['avatar']}
Offer: {lead['offer']}
Pain Point: {lead['pain']}
Desired Outcome: {lead['outcome']}

Use this structure:
1. Hero
2. Success State
3. Problem-Agitate-Transition
4. Value Stack
5. Social Proof
6. Transformation
7. Secondary CTA
8. Footer

Optimize for speed, clarity, and conversions.
"""

# =========================
# LOOM SCRIPT GENERATOR
# =========================
def build_loom_script(lead):
    return f"""
Write a casual Loom-style sales video script.

Lead name: {lead['name']}
Business: {lead['business']}
Pain point: {lead['pain']}
Offer: {lead['offer']}

Tone: friendly, confident, personalized.
Under 90 seconds.
"""

# =========================
# AI SMS COPY
# =========================
def build_sms_copy(lead):
    return f"""
Write a personalized SMS outreach message.

Recipient: {lead['name']}
Business: {lead['business']}
Pain point: {lead['pain']}

Goal: spark curiosity and reply.
Max 2 sentences.
"""

# =========================
# PRICING ENGINE
# =========================
def calculate_price(size, urgency, custom):
    base = {
        "solo": 2000,
        "smb": 4500,
        "enterprise": 10000
    }.get(size, 3000)

    if urgency == "high":
        base *= 1.2
    if custom == "yes":
        base *= 1.3

    return int(base)

# =========================
# NOTION SYNC
# =========================
def push_to_notion(lead, artifacts):
    url = "https://api.notion.com/v1/pages"
    headers = {
        "Authorization": f"Bearer {NOTION_API_KEY}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json"
    }

    data = {
        "parent": {"database_id": NOTION_DB_ID},
        "properties": {
            "Name": {"title": [{"text": {"content": lead["name"]}}]},
            "Business": {"rich_text": [{"text": {"content": lead["business"]}}]},
            "Price": {"number": artifacts["price"]},
            "AI Prompt": {"rich_text": [{"text": {"content": artifacts["web_prompt"][:2000]}}]},
            "Loom Script": {"rich_text": [{"text": {"content": artifacts["loom"][:2000]}}]},
            "SMS Copy": {"rich_text": [{"text": {"content": artifacts["sms"]}}]}
        }
    }

    requests.post(url, headers=headers, json=data)

# =========================
# INVOICE GENERATION
# =========================
def has_active_subscription(customer_id):
    subs = stripe.Subscription.list(customer=customer_id, status="active")
    return len(subs.data) > 0

def create_stripe_checkout(lead_name, price):
    session = stripe.checkout.Session.create(
        payment_method_types=["card"],
        line_items=[{
            "price_data": {
                "currency": "usd",
                "product_data": {
                    "name": f"{lead_name} – Custom Web App Build"
                },
                "unit_amount": price * 100,
            },
            "quantity": 1,
        }],
        mode="payment",
        success_url=f"{os.getenv('PUBLIC_CHECKOUT_DOMAIN')}/success",
        cancel_url=f"{os.getenv('PUBLIC_CHECKOUT_DOMAIN')}/cancel",
    )
    return session.url
    
def create_calendar_event(service, lead):
    event = {
        "summary": f"Sales Call – {lead['business']}",
        "description": "Discovery + walkthrough",
        "start": {"dateTime": "2026-02-10T14:00:00"},
        "end": {"dateTime": "2026-02-10T14:30:00"}
    }
    event = service.events().insert(calendarId="primary", body=event).execute()
    return event.get("htmlLink")

def generate_demo_site(lead):
    prompt = f"""
Generate a single-page HTML website for:
Business: {lead['business']}
Offer: {lead['offer']}
CTA: Book a Call

Use modern Tailwind-style layout.
"""
    html = ai_generate(prompt)
    file_name = f"demo_{lead['business']}.html"
    with open(file_name, "w") as f:
        f.write(html)
    return file_name


def has_website(lead):
    website = lead.get("website")
    if not website:
        return False
    if any(s in website.lower() for s in ["facebook.com", "instagram.com", "linkedin.com"]):
        return False
    return True


def generate_followup(business, step):
    prompt = f"Write follow-up #{step} for {business}. Keep it short."
    res = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )
    return res.choices[0].message.content.strip()

def classify_lead(lead):
    return "NO_WEBSITE" if not has_website(lead) else "HAS_WEBSITE"

def generate_no_website_message(lead):
    return f"""Hi {lead.get('name','there')},

I noticed your business in {lead.get('location')} doesn’t currently have a dedicated website.

We help businesses like yours attract more customers from Google, build trust online, and automate inquiries.

Would you be open to a quick walkthrough showing what this could look like for {lead.get('business_name')}?

Best,
{{Your Name}}
"""

def process_lead(lead):
    lead["lead_type"] = classify_lead(lead)
    if lead["lead_type"] == "NO_WEBSITE":
        lead["tailored_message"] = generate_no_website_message(lead)
    return lead

#example: https://www.google.com/maps/search/?api=1&query=Elite+Auto+Detailing+Dallas+TX
def generate_maps_link(business_name, location):
    query = f"{business_name} {location}"
    encoded = urllib.parse.quote_plus(query)
    return f"https://www.google.com/maps/search/?api=1&query={encoded}"

def score_lead(lead):
    score = 0
    if lead.get("lead_type") == "NO_WEBSITE":
        score += 40
    if lead.get("rating", 0) >= 4:
        score += 20
    if lead.get("reviews", 0) >= 20:
        score += 20
    return score

def fetch_linkedin_leads(phantom_id, api_key):
    url = f"https://api.phantombuster.com/api/v2/agents/fetch-output?id={phantom_id}"
    headers = {"X-Phantombuster-Key": api_key}
    res = requests.get(url, headers=headers).json()
    return res.get("data", [])

def scrape_x_leads(query):
    headers = {
        "Authorization": f"Bearer {os.getenv('X_BEARER_TOKEN')}"
    }
    params = {
        "query": query,
        "max_results": 10,
        "tweet.fields": "author_id,created_at"
    }
    res = requests.get(
        "https://api.twitter.com/2/tweets/search/recent",
        headers=headers,
        params=params
    ).json()
    return res.get("data", [])

def generate_x_dm(tweet_text, niche):
    prompt = f"""
    Write a short Twitter DM based on this tweet:

    "{tweet_text}"

    Niche: {niche['name']}
    Goal: Start a conversation, NOT pitch.
    """
    res = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )
    return res.choices[0].message.content.strip()

def generate_linkedin_dm(lead, niche, location):
    prompt = f"""
    Write a short, natural LinkedIn DM.

    Recipient:
    - Name: {lead.get('name')}
    - Role: {lead.get('title')}
    - Company: {lead.get('company')}
    - Location: {location}

    Niche: {niche['name']}
    Value Proposition: {niche['value_prop']}

    Rules:
    - Max 3 short sentences
    - NO selling
    - NO links
    - Sound human, casual, respectful
    - Goal: Start a conversation, not book a call

    End with a soft question.
    """

    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.7
    )

    return response.choices[0].message.content.strip()





# ---------------- SEND NOTIFICATIONS ---------------- #

# EMAIL
def notify_email(new_count, total_count):
    subject = "🚀 New Lead Added to Excel"
    body = f"""
    A new lead has been added to your Excel file.

    ➕ New entries: {new_count}
    📊 Total leads: {total_count}

    Check your Excel file for details.
    """

    message = Mail(
        from_email=os.getenv("NOTIFY_EMAIL_FROM"),
        to_emails=os.getenv("NOTIFY_EMAIL_TO"),
        subject=subject,
        plain_text_content=body
        # html_content=f"""
        # <strong>Business:</strong> {lead['Business Name']}<br>
        # <strong>Location:</strong> {lead['Location']}<br>
        # <strong>Lead Type:</strong> {lead['Lead Type']}<br>
        # <a href="{lead['Google Maps Link']}">View on Google Maps</a>
        # """
    )
    sg = SendGridAPIClient(os.getenv("SENDGRID_API_KEY"))
    sg.send(message)


def notify_email_lead(lead):
    message = Mail(
        from_email="alerts@yourdomain.com",
        to_emails="you@yourdomain.com",
        subject=f"🚀 New Lead: {lead['Business Name']}",
        # html_content=f"""
        # <strong>Business:</strong> {lead['Business Name']}<br>
        # <strong>Location:</strong> {lead['Location']}<br>
        # <strong>Lead Type:</strong> {lead['Lead Type']}<br>
        # <a href="{lead['Google Maps Link']}">View on Google Maps</a>
        # """
    )
    sg = SendGridAPIClient(os.getenv("SENDGRID_API_KEY"))
    sg.send(message)
    
# TELEGRAM
def notify_telegram(lead, new_count, total_count):
    message = (
        f"🚀 *New Lead Added!*\n\n"
        f"➕ New entries: {new_count}\n"
        f"📊 Total leads: {total_count}"
        f"*Business:* {lead['Business Name']}\n"
        f"*Location:* {lead['Location']}\n"
        f"*Type:* {lead['Lead Type']}\n"
        f"*Score:* {lead['Lead Score']}\n"
        f"[Open in Google Maps]({lead['Google Maps Link']})"
    )

    url = f"https://api.telegram.org/bot{os.getenv('TELEGRAM_BOT_TOKEN')}/sendMessage"
    payload = {
        "chat_id": os.getenv("TELEGRAM_CHAT_ID"),
        "text": message,
        "parse_mode": "Markdown"
    }

    requests.post(url, json=payload)


def notify_telegram_lead(lead):
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")

    text = (
        f"🚀 *New Lead*\n\n"
        f"*Business:* {lead['Business Name']}\n"
        f"*Location:* {lead['Location']}\n"
        f"*Type:* {lead['Lead Type']}\n"
        f"[Open in Google Maps]({lead['Google Maps Link']})"
    )

    requests.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        json={
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "Markdown"
        }
    )
def notify_sms(lead):
    client = Client(
        os.getenv("TWILIO_ACCOUNT_SID"),
        os.getenv("TWILIO_AUTH_TOKEN")
    )
    client.messages.create(
        body=f"New Lead: {lead['Business Name']} - {lead['Google Maps Link']}",
        from_=os.getenv("TWILIO_PHONE_NUMBER"),
        to=os.getenv("ALERT_PHONE_NUMBER")
    )

def generate_ai_sms(lead):
    prompt = f"""Write a concise, friendly SMS for a business owner.
Business: {lead['Business Name']}
Location: {lead['Location']}
Context: Business does not have a website.
Tone: Helpful, non-salesy."""
    resp = openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.choices[0].message.content.strip()

def send_sms(lead):
    body = generate_ai_sms(lead)
    client = Client(os.getenv("TWILIO_ACCOUNT_SID"), os.getenv("TWILIO_AUTH_TOKEN"))
    client.messages.create(
        body=body,
        from_=os.getenv("TWILIO_PHONE_NUMBER"),
        to=os.getenv("ALERT_PHONE_NUMBER")
    )
    
def notify_all(lead):
    notify_email(lead)
    notify_telegram(lead)
    notify_sms(lead)


# ---------------- CRM UTILITY ---------------- #
# EXPORT SHEET TO EXCEL
def export_sheet_to_excel(spreadsheet_id, sheet_name="Sheet1"):
    creds = Credentials.from_service_account_file(
        "service_account.json",
        scopes=["https://www.googleapis.com/auth/drive"]
    )

    drive_service = build("drive", "v3", credentials=creds)

    today = datetime.date.today().isoformat()
    file_name = f"leads_{today}.xlsx"

    request = drive_service.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with open(file_name, "wb") as f:
        f.write(request.execute())

    print(f"Exported Excel file: {file_name}")
    
# NOTION SYNC
def sync_to_notion(lead):
    headers = {
        "Authorization": f"Bearer {os.getenv('NOTION_API_KEY')}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json"
    }
    data = {
        "parent": {"database_id": os.getenv("NOTION_DATABASE_ID")},
        "properties": {
            "Business Name": {"title": [{"text": {"content": lead['Business Name']}}]},
            "Location": {"rich_text": [{"text": {"content": lead['Location']}}]},
            "Lead Type": {"select": {"name": lead['Lead Type']}},
            "Google Maps": {"url": lead['Google Maps Link']}
        }
    }
    requests.post("https://api.notion.com/v1/pages", headers=headers, json=data)
    
# EXCEL NEW-LEAD DETECTOR
def load_last_row_count():
    if not os.path.exists(STATE_FILE):
        return 0
    with open(STATE_FILE, "r") as f:
        return json.load(f).get("row_count", 0)

def save_row_count(count):
    with open(STATE_FILE, "w") as f:
        json.dump({"row_count": count}, f)

def get_excel_row_count(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    return ws.max_row - 1  # exclude header
    
# GOOGLE SHEETS NEW-LEAD DETECTOR
def load_state():
    if os.path.exists(STATE_FILE):
        return json.load(open(STATE_FILE))
    return {"last_row": 1}

def save_state(row):
    json.dump({"last_row": row}, open(STATE_FILE, "w"))

def get_new_leads(sheet):
    state = load_state()
    last_row = state["last_row"]

    all_rows = sheet.get_all_records()
    current_row_count = len(all_rows) + 1  # header row

    if current_row_count <= last_row:
        return []

    new_leads = all_rows[last_row - 1 :]
    save_state(current_row_count)

    return new_leads
def process_new_sheet_entries(sheet):
    new_leads = get_new_leads(sheet)

    for lead in new_leads:
        notify_email(lead)
        notify_telegram(lead)


# ---------------- UTILS ---------------- #
def lead_hash(email):
    return hashlib.md5(email.encode()).hexdigest()

def already_contacted(email):
    records = sheet.get_all_records()
    hashes = [lead_hash(r["Email"]) for r in records if r["Email"]]
    return lead_hash(email) in hashes

# ---------------- GOOGLE MAPS SCRAPER ---------------- #
def scrape_google_maps(query, location):
    params = {
        "engine": "google_maps",
        "q": query,
        "location": location,
        "api_key": SERPAPI_KEY
    }
    res = requests.get("https://serpapi.com/search", params=params).json()
    return res.get("local_results", [])[:limit]

# ---------------- AI EMAIL GENERATOR ---------------- #
def generate_email(business):
    prompt = f"""
    Write a personalized cold outreach email to:
    Business: {business['title']}
    Website: {business.get('website')}
    Location: {business.get('address')}
    Goal: Book a short sales call
    Tone: Professional, direct, friendly
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content.strip()

# ---------------- SEND EMAIL ---------------- #
def send_email(to_email, subject, content):
    message = Mail(
        from_email="you@yourdomain.com",
        to_emails=to_email,
        subject=subject,
        html_content=content
    )
    sg = SendGridAPIClient(SENDGRID_KEY)
    sg.send(message)

# ---------------- CREATE CALENDAR EVENT ---------------- #
def book_call(business_name, email):
    event = {
        "summary": f"Intro Call - {business_name}",
        "description": "Automated sales call booking",
        "start": {
            "dateTime": (datetime.datetime.utcnow() + datetime.timedelta(days=2)).isoformat() + "Z"
        },
        "end": {
            "dateTime": (datetime.datetime.utcnow() + datetime.timedelta(days=2, minutes=30)).isoformat() + "Z"
        },
        "attendees": [{"email": email}],
        "conferenceData": {
            "createRequest": {
                "requestId": "meet123",
                "conferenceSolutionKey": {"type": "hangoutsMeet"}
            }
        }
    }

    event = calendar.events().insert(
        calendarId=CALENDAR_ID,
        body=event,
        conferenceDataVersion=1
    ).execute()

    return event["hangoutLink"]

# ---------------- FOLLOW-UP GENERATOR ---------------- #
def generate_follow_up(business_name, step):
    prompt = f"""
    Write follow-up email #{step} for {business_name}.
    Keep it short and polite.
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content.strip()


# =========================
# MAIN PIPELINE
# =========================

# ---------------- MAIN AGENT ---------------- #

def run_v7():
    rows = sheet.get_all_records()

    for i, row in enumerate(rows, start=2):
        lead = {
            "name": row["Name"],
            "business": row["Business"],
            "industry": row["Industry"],
            "avatar": row["Avatar"],
            "offer": row["Offer"],
            "pain": row["Pain Point"],
            "outcome": row["Desired Outcome"],
            "size": row["Business Size"],
            "urgency": row["Urgency"],
            "custom": row["Custom Build"]
        }

        web_prompt = ai_generate(build_web_app_prompt(lead))
        loom_script = ai_generate(build_loom_script(lead))
        sms_copy = ai_generate(build_sms_copy(lead))
        price = calculate_price(lead["size"], lead["urgency"], lead["custom"])
        checkout = create_stripe_checkout(lead["business"], price)
        proposal = generate_proposal_pdf(lead, price)
        demo_site = generate_demo_site(lead)

        # Write back to Google Sheet
        sheet.update(f"H{i}", web_prompt)
        sheet.update(f"I{i}", loom_script)
        sheet.update(f"J{i}", sms_copy)
        sheet.update(f"K{i}", price)
        sheet.update(f"L{i}", checkout)
        sheet.update(f"M{i}", proposal)
        sheet.update(f"N{i}", demo_site)

        # Push to Notion
        push_to_notion(lead, {
            "web_prompt": web_prompt,
            "loom": loom_script,
            "sms": sms_copy,
            "price": price
        })

        print(f"✅ Processed lead: {lead['name']}")
def run_agent():
    today = str(datetime.date.today())

    sources = ["google_maps", "linkedin", "x"]

    for source in sources:
        if source == "google_maps":
            ...
        elif source == "linkedin":
            ...
        elif source == "x":
            ...

    # 10 niches × 20 cities × 10 leads/day = 2,000 new leads/day
    for niche in CONFIG["niches"]:
        for location in CONFIG["locations"]:
            leads = scrape_maps(
                niche["search_query"],
                location,
                CONFIG["daily_limit_per_combo"]
            )
            linkedin_leads = fetch_linkedin_leads(
                phantom_id=os.getenv("PHANTOM_ID"),
                api_key=os.getenv("PHANTOMBUSTER_API_KEY")
            )

            for lead in leads:
                email = lead.get("email")
                if not email or already_contacted(email):
                    continue

                initial = generate_email(lead, niche, location)
                follow1 = generate_followup(lead["title"], 1)
                follow2 = generate_followup(lead["title"], 2)

                calendar_link = book_call(lead["title"], email)

                # send_email(email, "Quick question", initial)

                sheet.append_row([
                    niche["name"],
                    location,
                    lead["title"],
                    lead.get("website"),
                    email,
                    lead.get("phone"),
                    initial,
                    follow1,
                    follow2,
                    calendar_link, #" ",
                    "Contacted",
                    today,
                    "Google Maps"
                ])
            
            for lead in linkedin_leads:
                sheet.append_row([
                    niche["name"],
                    lead["location"],  #lead.get("location")
                    lead["company"],   #lead.get("company")
                    "",                #website
                    lead.get("email"), #email address
                    "",                #phone number
                    initial_email,     #initial email
                    follow1,           #follow-up 1
                    follow2,           #follow-up 2
                    "",                #calendar link
                    "Queued",          #status
                    today,             #last contacted
                    "LinkedIn",        #lead source
                    lead["profileUrl"],#lead.get("profileUrl")
                    lead["name"],      #lead.get("name")
                    "DM",              #outreach type
                    "",
                    linkedin_dm,
                    "Not Sent"
                ])
    # leads = scrape_google_maps("Marketing Agency", "New York")

    # for lead in leads:
    #     email = lead.get("email")
    #     if not email:
    #         continue

    #     initial_email = generate_email(lead)
    #     follow1 = generate_follow_up(lead["title"], 1)
    #     follow2 = generate_follow_up(lead["title"], 2)

    #     calendar_link = book_call(lead["title"], email)

    #     send_email(email, "Quick question", initial_email)

    #     sheet.append_row([
    #         lead["title"],
    #         lead.get("website"),
    #         email,
    #         lead.get("phone"),
    #         lead.get("address"),
    #         initial_email,
    #         follow1,
    #         follow2,
    #         calendar_link,
    #         "Contacted",
    #         str(datetime.date.today())
    #     ])

if __name__ == "__main__":
    run_agent()
    # export_sheet_to_excel("YOUR_SPREADSHEET_ID")
    # Export Excel
    excel_file = export_sheet_to_excel("YOUR_SPREADSHEET_ID")
    cloud_sheet_file = "cloud_sheet.xlsx"
    previous_count = load_last_row_count()
    current_count = get_excel_row_count(excel_file)

    if current_count > previous_count:
        new_entries = current_count - previous_count

        notify_email(new_entries, current_count)
        notify_telegram(new_entries, current_count)

    save_row_count(current_count)

# Stripe Example usage:
# if not has_active_subscription(client["stripe_customer_id"]):
#     return
