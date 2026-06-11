import json
import requests
import time
from typing import Dict, List, Any
import os
from openpyxl import load_workbook, Workbook
import os
import json
import datetime
import time
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from apify_client import ApifyClient
from langchain_openai import ChatOpenAI
from langchain.core.prompts import PromptTemplate
from langchain.core.output_parsers import JsonOutputParser
from pydantic import BaseModel, Field

# 1. GUARDIAN RUN TRACKER
TRACKER_FILE = "daily_quota_tracker.txt"

# SCOPES AND AGENT STRUCTURE DEFINITION
GOOGLE_SCOPES = [
    "https://googleapis.com",
    "https://googleapis.com"
]
# ==========================================
# REAL ESTATE ETL ENGINE (ZILLOW API)
# ==========================================



class InvestmentEvaluation(BaseModel):
    score: int = Field(description="An investment readiness score from 1 to 100.")
    rationale: str = Field(description="A 1-2 sentence summary explaining the score.")

def load_pipeline_config(config_path="pipeline_config.json"):
    """Loads configuration metadata cleanly from the local JSON file."""
    with open(config_path, "r") as f:
        return json.load(f)

# 2. TARGET RUNNER FOR A SPECIFIC LOCATION
def process_location(client, ai_pipeline, location, country, settings):
    """Scrapes and returns an AI-evaluated dataframe for a specific market block."""
    print(f"🌍 Processing: {location} ({country})...")
    
    actor_input = {
      "search": location,
      "type": "FOR_SALE",
      "maxPages": settings["max_pages_per_run"],
      "resultsPerPage": settings["results_per_page"]
    }
    
    try:
        run = client.actor("scrapier/zillow-search-scraper").call(
            run_input=actor_input, 
            timeout_secs=settings["api_timeout_seconds"]
        )
        items = list(client.dataset(run["defaultDatasetId"]).iterate_items())
        
        if not items:
            print(f"⚠️ No property matches found for {location}.")
            return None
            
        raw_df = pd.DataFrame(items)
        
        # Standardize property categories
        type_mapping = {
            'SINGLE_FAMILY': 'Single-Family', 'CONDO': 'Condo', 'TOWNHOUSE': 'Townhome',
            'MULTI_FAMILY': 'Multi-Tenant', 'LOT': 'Lots/Land', 'MANUFACTURED': 'Single-Family',
            'APARTMENT': 'Multi-Tenant'
        }
        if 'homeType' not in raw_df.columns:
            raw_df['homeType'] = 'SINGLE_FAMILY'
        raw_df['Category'] = raw_df['homeType'].map(type_mapping).fillna('Single-Family')
        
        # Clean data structures
        cols_to_keep = ['address', 'price', 'bedrooms', 'bathrooms', 'livingArea', 'Category', 'url']
        existing_cols = [c for c in cols_to_keep if c in raw_df.columns]
        df = raw_df[existing_cols].copy().dropna(subset=['price'])
        
        df.rename(columns={
            'address': 'Address', 'price': 'Price', 'bedrooms': 'Beds', 
            'bathrooms': 'Baths', 'livingArea': 'SqFt', 'url': 'Zillow Link'
        }, inplace=True)

        df['Beds'] = df['Beds'].fillna("N/A")
        df['Baths'] = df['Baths'].fillna("N/A")
        df['SqFt'] = df['SqFt'].fillna("N/A")

        # Run records sequentially through AI evaluation
        ai_scores, ai_rationales = [], []
        for _, row in df.iterrows():
            try:
                res = ai_pipeline.invoke({
                    "category": row['Category'], "price": row['Price'],
                    "beds": row['Beds'], "baths": row['Baths'], "sqft": row['SqFt'],
                    "location": f"{location}, {country}"
                })
                ai_scores.append(res.get("score", 50))
                ai_rationales.append(res.get("rationale", "Evaluation skipped."))
            except:
                ai_scores.append(50)
                ai_rationales.append("Failed to evaluate parameters.")
        
        df['Investment Score'] = ai_scores
        df['AI Analysis Summary'] = ai_rationales
        
        # Add tracking metadata fields
        df['Country'] = country
        df['Search Location'] = location
        df['Date Discovered'] = datetime.date.today().isoformat()
        
        return df
        
    except Exception as e:
        print(f"❌ Error compiling data for {location}: {str(e)}")
        return None

# 3. CORE MANAGEMENT CONTROLLER
def run_global_pipeline():
    config = load_pipeline_config()
    
    apify_token = os.getenv("APIFY_API_TOKEN", "your_apify_token_here")
    openai_key = os.getenv("OPENAI_API_KEY", "your_openai_key_here")
    
    client = ApifyClient(apify_token)
    
    # Initialize LangChain tools
    llm = ChatOpenAI(temperature=0, model="gpt-4o", openai_api_key=openai_key)
    parser = JsonOutputParser(pydantic_object=InvestmentEvaluation)
    prompt = PromptTemplate(
        template="You are a real estate valuation algorithm.\n{format_instructions}\nAnalyze this property:\nCategory: {category}\nPrice: ${price}\nLayout: {beds} Bed / {baths} Bath, SqFt: {sqft}\nLocation Context: {location}\nReturn a calculated score (1-100) and an absolute clear investment summary.",
        input_variables=["category", "price", "beds", "baths", "sqft", "location"],
        partial_variables={"format_instructions": parser.get_format_instructions()}
    )
    ai_pipeline = prompt | llm | parser

    all_market_dfs = []
    
    # Loop over everything specified in the JSON file
    for market in config["target_markets"]:
        country = market["country"]
        for location in market["locations"]:
            market_df = process_location(client, ai_pipeline, location, country, config["settings"])
            if market_df is not None:
                all_market_dfs.append(market_df)
            time.sleep(2) # Brief cooldown pause to maintain connection health

    if not all_market_dfs:
        print("❌ No items retrieved across any markets.")
        return

    # Merge all scraped data together into one unified table structure
    final_combined_df = pd.concat(all_market_dfs, ignore_index=True)
    
    # Organize columns cleanly for Google Sheets
    column_layout = ['Date Discovered', 'Country', 'Search Location', 'Address', 'Category', 'Price', 'Beds', 'Baths', 'SqFt', 'Investment Score', 'AI Analysis Summary', 'Zillow Link']
    final_combined_df = final_combined_df[column_layout].sort_values(by='Investment Score', ascending=False)

    # Stream results directly up to Google Sheet Workspace
    print("📁 Connecting to Google Cloud Platform...")
    creds = Credentials.from_service_account_file("google_creds.json", scopes=GOOGLE_SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(config["spreadsheet_id"])
    
    try:
        worksheet = sh.worksheet("REAL_ESTATE_LEADS")
    except gspread.exceptions.WorksheetNotFound:
        worksheet = sh.add_worksheet(title="REAL_ESTATE_LEADS", rows="1000", cols="12")

    # Fetch existing sheet entries to check if we need to write fresh headers
    existing_records = worksheet.get_all_values()
    
    data_rows = final_combined_df.values.tolist()
    
    if not existing_records:
        # Sheet is empty: write header + data rows together
        worksheet.update(range_name="A1", values=[column_layout] + data_rows)
    else:
        # Sheet has data: append new runs cleanly to the bottom rows
        worksheet.append_rows(data_rows)
        
    print(f"🎉 Success! Combined listings successfully updated in your REAL_ESTATE_LEADS workspace tab.")

def check_and_update_daily_quota() -> bool:
    """Blocks multi-execution to preserve Apify and OpenAI credits."""
    today_str = datetime.date.today().isoformat()
    if os.path.exists(TRACKER_FILE):
        with open(TRACKER_FILE, "r") as file:
            if file.read().strip() == today_str:
                return False
    with open(TRACKER_FILE, "w") as file:
        file.write(today_str)
    return True


# 2. DATASET INGESTION & ANALYSIS PIPELINE
def run_zillow_google_sheets_pipeline(spreadsheet_id: str, location: str):
    print(f"🚀 Initializing Zillow extraction pipeline for: {location}...")
    
    actor_input = {
        "search": location,
        "type": "FOR_SALE",
        "maxPages": 1,
        "resultsPerPage": 8  # Optimized chunk size to conserve processing limits
    }
    
    apify_token = os.getenv("APIFY_API_TOKEN", "your_apify_token_here")
    openai_key = os.getenv("OPENAI_API_KEY", "your_openai_key_here")
    
    client = ApifyClient(apify_token)
    
    try:
        # Step A: Run the Web Scraper Task
        run = client.actor("scrapier/zillow-search-scraper").call(run_input=actor_input, timeout_secs=120)
        items = list(client.dataset(run["defaultDatasetId"]).iterate_items())
        
        if not items:
            print("❌ No property data matches found.")
            return
            
        raw_df = pd.DataFrame(items)
        
        # Step B: Standardize Asset Classifications
        type_mapping = {
            'SINGLE_FAMILY': 'Single-Family', 'CONDO': 'Condo', 'TOWNHOUSE': 'Townhome',
            'MULTI_FAMILY': 'Multi-Tenant', 'LOT': 'Lots/Land', 'MANUFACTURED': 'Single-Family',
            'APARTMENT': 'Multi-Tenant'
        }
        if 'homeType' not in raw_df.columns:
            raw_df['homeType'] = 'SINGLE_FAMILY'
        raw_df['Category'] = raw_df['homeType'].map(type_mapping).fillna('Single-Family')
        
        # Keep and sanitize core layout fields
        cols_to_keep = ['address', 'price', 'bedrooms', 'bathrooms', 'livingArea', 'Category', 'url']
        existing_cols = [c for c in cols_to_keep if c in raw_df.columns]
        df = raw_df[existing_cols].copy().dropna(subset=['price'])
        
        df.rename(columns={
            'address': 'Address', 'price': 'Price', 'bedrooms': 'Beds', 
            'bathrooms': 'Baths', 'livingArea': 'SqFt', 'url': 'Zillow Link'
        }, inplace=True)

        # Fill potential empty fields with standard text so Google Sheets doesn't break
        df['Beds'] = df['Beds'].fillna("N/A")
        df['Baths'] = df['Baths'].fillna("N/A")
        df['SqFt'] = df['SqFt'].fillna("N/A")

        # Step C: Initialize the AI Investment Evaluator
        llm = ChatOpenAI(temperature=0, model="gpt-4o", openai_api_key=openai_key)
        parser = JsonOutputParser(pydantic_object=InvestmentEvaluation)
        
        prompt = PromptTemplate(
            template="You are a real estate valuation algorithm.\n{format_instructions}\nAnalyze this property:\nCategory: {category}\nPrice: ${price}\nLayout: {beds} Bed / {baths} Bath, SqFt: {sqft}\nLocation Context: {location}\nReturn a calculated score (1-100) and an absolute clear investment summary.",
            input_variables=["category", "price", "beds", "baths", "sqft", "location"],
            partial_variables={"format_instructions": parser.get_format_instructions()}
        )
        
        ai_pipeline = prompt | llm | parser
        
        ai_scores, ai_rationales = [], []
        print("🧠 Invoking AI scoring patterns...")
        for _, row in df.iterrows():
            try:
                res = ai_pipeline.invoke({
                    "category": row['Category'], "price": row['Price'],
                    "beds": row['Beds'], "baths": row['Baths'], "sqft": row['SqFt'],
                    "location": location
                })
                ai_scores.append(res.get("score", 50))
                ai_rationales.append(res.get("rationale", "Evaluation skipped."))
            except:
                ai_scores.append(50)
                ai_rationales.append("Error parsing record analysis parameters.")
        
        df['Investment Score'] = ai_scores
        df['AI Analysis Summary'] = ai_rationales
        
        # Order and sort records by investment yield probability
        final_order = ['Address', 'Category', 'Price', 'Beds', 'Baths', 'SqFt', 'Investment Score', 'AI Analysis Summary', 'Zillow Link']
        df = df[[c for c in final_order if c in df.columns]].sort_values(by='Investment Score', ascending=False)

        # Step D: Connect and stream directly into Google Sheets
        print("📁 Streaming transformed listings into Google Sheets cloud...")
        creds = Credentials.from_service_account_file("google_creds.json", scopes=GOOGLE_SCOPES)
        gc = gspread.authorize(creds)
        
        # Open your master spreadsheet file
        sh = gc.open_by_key(spreadsheet_id)
        
        # Dynamically find or build the specific REAL_ESTATE_LEADS table
        try:
            worksheet = sh.worksheet("REAL_ESTATE_LEADS")
        except gspread.exceptions.WorksheetNotFound:
            print("⚠️ 'REAL_ESTATE_LEADS' tab not found. Building a new table workspace...")
            worksheet = sh.add_worksheet(title="REAL_ESTATE_LEADS", rows="100", cols="10")
            
        # Format dataset content data directly into pure lists for Google API transit
        header = df.columns.tolist()
        data_rows = df.values.tolist()
        
        # Clear any existing rows inside the sheet and rewrite clean data
        worksheet.clear()
        worksheet.update(range_name="A1", values=[header] + data_rows)
        print("🎉 Cloud Synchronization Complete! Data is now live.")
        
    except Exception as e:
        print(f"❌ Execution failed: {str(e)}")



def export_real_estate_to_excel(categorized_leads: dict, filename: str = "leads.xlsx"):
    """
    Appends or creates a dedicated 'Real Estate Leads' worksheet 
    inside the main workbook, tracking investment strategy types.
    """
    sheet_name = "Real Estate Leads"
    
    # 1. Initialize or Load Workbook
    if os.path.exists(filename):
        wb = load_workbook(filename)
        # Grab existing sheet or create it fresh if it doesn't exist
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # 2. Setup Headers if the sheet is completely empty
    headers = [
        "Strategy", "ZPID", "Address", "Price", 
        "Property Type", "Beds", "Baths", 
        "Days on Market", "Zestimate", "URL", "Notes"
    ]
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(headers)

    # 3. Flatten and append the categorized dictionaries
    row_count = 0
    for strategy, properties in categorized_leads.items():
        for prop in properties:
            row_data = [
                strategy.replace("_", " ").upper(),
                prop.get("zpid", "N/A"),
                prop.get("address", "N/A"),
                prop.get("price", 0),
                prop.get("property_type", "N/A"),
                prop.get("beds", 0),
                prop.get("baths", 0),
                prop.get("days_on_market", -1),
                prop.get("zestimate", "N/A"),
                prop.get("url", "N/A"),
                prop.get("wholesale_trigger", "")  # Holds keyword/DOM trigger context
            ]
            ws.append(row_data)
            row_count += 1

    # 4. Save workbook state back safely
    wb.save(filename)
    print(f"[INFO] Successfully exported {row_count} real estate leads to sheet: '{sheet_name}' inside {filename}")

def load_config(config_path: str = "config.json") -> Dict[str, Any]:
    """Loads configuration schema safely."""
    try:
        with open(config_path, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"[ERROR] Failed to read config file: {e}")
        return {}

def fetch_zillow_properties(location: str, config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Polls RapidAPI Zillow endpoint using search parameters.
    Handles primary list-view retrieval.
    """
    api_key = config.get("rapidapi_key")
    api_host = config.get("rapidapi_host", "zillow56.p.rapidapi.com")
    
    if not api_key or api_key == "YOUR_RAPIDAPI_KEY_HERE":
        print("[ERROR] Valid RapidAPI key missing from config.json")
        return []

    url = f"https://{api_host}/search"
    headers = {
        "X-RapidAPI-Key": api_key,
        "X-RapidAPI-Host": api_host
    }
    
    # Generic parameter injection covering core location targets
    querystring = {
        "location": location, 
        "status": "FOR_SALE",
        "page": "1"
    }

    try:
        print(f"[INFO] Extracting Zillow raw data for location: {location}...")
        response = requests.get(url, headers=headers, params=querystring, timeout=15)
        
        if response.status_code == 429:
            print("[WARN] Rate-limited (429). Exponential backoff triggered...")
            time.sleep(5)
            return []
            
        if response.status_code != 200:
            print(f"[ERROR] API failed with status {response.status_code}: {response.text}")
            return []
            
        data = response.json()
        return data.get("results", [])
        
    except requests.exceptions.RequestException as e:
        print(f"[CRITICAL] Network drop or timeout during API hit: {e}")
        return []

def transform_and_evaluate(properties: List[Dict[str, Any]], re_config: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Runs extraction, schema transformation, and evaluates matches across 
    investment, wholesaling, and target premium buy strategies.
    """
    categorized_leads = {
        "investment_buy_hold": [],
        "wholesaling": [],
        "buying_premium": []
    }
    
    strategies = re_config.get("strategies", {})
    allowed_types = re_config.get("property_types", ["SINGLE_FAMILY", "MULTI_FAMILY", "LAND"])

    for prop in properties:
        # Standardize properties layout safely from variable API schemas
        home_type = prop.get("homeType", "UNKNOWN")
        price = prop.get("price", 0)
        beds = prop.get("bedrooms", 0)
        baths = prop.get("bathrooms", 0)
        days_on_zillow = prop.get("daysOnZillow", -1)
        description = prop.get("description", "").lower()
        zestimate = prop.get("zestimate")
        year_built = prop.get("yearBuilt", 0)
        zpid = prop.get("zpid")

        # Global Baseline Type Constraint Check
        if home_type not in allowed_types and len(allowed_types) > 0:
            continue

        # Normalized Payload for clean processing down-funnel
        normalized_lead = {
            "zpid": zpid,
            "address": prop.get("address", "N/A"),
            "price": price,
            "property_type": home_type,
            "beds": beds,
            "baths": baths,
            "days_on_market": days_on_zillow,
            "zestimate": zestimate,
            "url": f"https://www.zillow.com/homedetails/{zpid}_zpid/" if zpid else "N/A"
        }

        # 1. Strategy Evaluation: Investment (Buy & Hold)
        ib_cfg = strategies.get("investment_buy_hold", {})
        if (price <= ib_cfg.get("max_price", 9999999) and 
            beds >= ib_cfg.get("min_bedrooms", 0) and 
            baths >= ib_cfg.get("min_bathrooms", 0)):
            if not ib_cfg.get("requires_zestimate") or zestimate is not None:
                categorized_leads["investment_buy_hold"].append(normalized_lead)

        # 2. Strategy Evaluation: Wholesaling (Looking for deep distress metrics)
        ws_cfg = strategies.get("wholesaling", {})
        if price <= ws_cfg.get("max_price", 9999999):
            # Keyword triggers for deep value-add properties
            has_keyword = any(kw in description for kw in ws_cfg.get("keywords", []))
            # Stale listings often reveal highly motivated sellers
            is_stale = days_on_zillow >= ws_cfg.get("max_days_on_zillow", 90)
            
            if has_keyword or is_stale:
                normalized_lead["wholesale_trigger"] = "Keyword Match" if has_keyword else "High Days on Market"
                categorized_leads["wholesaling"].append(normalized_lead)

        # 3. Strategy Evaluation: Premium Buy Criteria
        bp_cfg = strategies.get("buying_premium", {})
        if (bp_cfg.get("min_price", 0) <= price <= bp_cfg.get("max_price", 9999999) and 
            year_built >= bp_cfg.get("min_year_built", 0)):
            categorized_leads["buying_premium"].append(normalized_lead)

    return categorized_leads

def run_real_estate_pipeline():
    """Main execution block hook for pipeline orchestrator."""
    config = load_config()
    re_config = config.get("real_estate_config", {})
    
    if not re_config:
        print("[WARN] No real_estate_config blocks found. Execution halted.")
        return

    all_extracted_leads = {
        "investment_buy_hold": [],
        "wholesaling": [],
        "buying_premium": []
    }

    # Iterate structural parameters built inside current config loops
    for location in config.get("locations", []):
        raw_properties = fetch_zillow_properties(location, config)
        if not raw_properties:
            continue
            
        segmented = transform_and_evaluate(raw_properties, re_config)
        
        for strategy in all_extracted_leads.keys():
            all_extracted_leads[strategy].extend(segmented[strategy])
            
        # Throttling to prevent API execution blockades across rapid loops
        time.sleep(1.5)

    # Output analytical matrix breakdown
    # print("\n================== PIPELINE SUMMARY ==================")
    # for strategy, leads in all_extracted_leads.items():
    #     print(f"Strategy [{strategy.upper()}]: Found {len(leads)} viable properties matches.")
    # print("======================================================")
    
    # --- NEW EXPORT TRIGGER ---
    if config.get("excel_export", True):
        export_real_estate_to_excel(all_extracted_leads, filename="leads.xlsx")
    
    # Ready for integration into your downstream excel_export or messaging modules.
    return all_extracted_leads



# 3. CORE REAL ESTATE RUN PIPELINE
def run_zillow_investment_pipeline(location: str):
    if not check_and_update_daily_quota():
        print("⛔ Quota Limit Triggered: Script already executed today.")
        return

    print(f"🚀 Pulling live listings for: {location}...")
    
    # Platform limits applied to stay strictly inside the free tier
    actor_input = {
        "search": location,
        "type": "FOR_SALE",
        "maxPages": 1,
        "resultsPerPage": 10
    }
    
    apify_token = os.getenv("APIFY_API_TOKEN", "your_apify_token_here")
    openai_key = os.getenv("OPENAI_API_KEY", "your_openai_key_here")
    
    client = ApifyClient(apify_token)
    
    try:
        # Fetching raw structured property data
        run = client.actor("scrapier/zillow-search-scraper").call(run_input=actor_input, timeout_secs=120)
        items = list(client.dataset(run["defaultDatasetId"]).iterate_items())
        
        if not items:
            print("❌ No items retrieved from the search grid.")
            return
            
        raw_df = pd.DataFrame(items)
        
        # 4. ZILLOW INTERNAL CATEGORY MAPPING
        # Maps Zillow's internal 'homeType' strings cleanly into your exact Excel categories
        type_mapping = {
            'SINGLE_FAMILY': 'Single-Family',
            'CONDO': 'Condo',
            'TOWNHOUSE': 'Townhome',
            'MULTI_FAMILY': 'Multi-Tenant',
            'LOT': 'Lots/Land',
            'MANUFACTURED': 'Single-Family',
            'APARTMENT': 'Multi-Tenant'
        }
        
        # Ensure fallback column exists if homeType is missing
        if 'homeType' not in raw_df.columns:
            raw_df['homeType'] = 'SINGLE_FAMILY'
            
        raw_df['Category'] = raw_df['homeType'].map(type_mapping).fillna('Single-Family')
        
        # Keep clean, vital listing data structures
        cols_to_keep = ['address', 'price', 'bedrooms', 'bathrooms', 'livingArea', 'Category', 'url']
        existing_cols = [c for c in cols_to_keep if c in raw_df.columns]
        df = raw_df[existing_cols].copy().dropna(subset=['price'])
        
        # Rename layout fields for presentation
        df.rename(columns={
            'address': 'Address', 
            'price': 'Price', 
            'bedrooms': 'Beds', 
            'bathrooms': 'Baths', 
            'livingArea': 'SqFt', 
            'url': 'Zillow Link'
        }, inplace=True)

        # 5. INITIALIZE THE STRUCTURED AI SCORING MODEL
        llm = ChatOpenAI(temperature=0, model="gpt-4o", openai_api_key=openai_key)
        parser = JsonOutputParser(pydantic_object=InvestmentEvaluation)
        
        prompt = PromptTemplate(
            template="You are an expert real estate data analyst.\n{format_instructions}\nEvaluate this property:\nCategory: {category}\nPrice: ${price}\nBeds: {beds}, Baths: {baths}, SqFt: {sqft}\nLocation Context: {location}\nProvide a score (1-100) assessing its value proposition (e.g. high square footage for the price, logical room counts, or developmental land viability) and a brief rationale.",
            input_variables=["category", "price", "beds", "baths", "sqft", "location"],
            partial_variables={"format_instructions": parser.get_format_instructions()}
        )
        
        ai_pipeline = prompt | llm | parser
        
        # Storage lists for our new AI-generated columns
        ai_scores = []
        ai_rationales = []
        
        print("🧠 Analyzing deals using AI Engine...")
        for _, row in df.iterrows():
            try:
                # Format variables safely for the LLM
                res = ai_pipeline.invoke({
                    "category": row['Category'],
                    "price": row['Price'],
                    "beds": row.get('Beds', 'N/A'),
                    "baths": row.get('Baths', 'N/A'),
                    "sqft": row.get('SqFt', 'N/A'),
                    "location": location
                })
                ai_scores.append(res.get("score", 50))
                ai_rationales.append(res.get("rationale", "No analysis generated."))
            except Exception as ai_err:
                ai_scores.append(50)
                ai_rationales.append("AI processing encountered an omission error.")
        
        # Append the new scoring metrics to our spreadsheet pipeline
        df['Investment Score'] = ai_scores
        df['AI Analysis Summary'] = ai_rationales
        
        # Reorder columns so the scores are front and center
        final_order = ['Address', 'Category', 'Price', 'Beds', 'Baths', 'SqFt', 'Investment Score', 'AI Analysis Summary', 'Zillow Link']
        df = df[[c for c in final_order if c in df.columns]]
        
        # Sort listings so the highest scoring investment opportunities float to the top
        df = df.sort_values(by='Investment Score', ascending=False)

        # 6. EXPORTING CRADLE TO EXCEL
        output_file = "zillow_investment_analysis.xlsx"
        df.to_excel(output_file, index=False)
        print(f"🎉 Success! Spreadsheet built and saved to: {output_file}")
        
    except Exception as e:
        print(f"❌ Structural error encountered: {str(e)}")

if __name__ == "__main__":
    # Test execution for your chosen market
    run_zillow_investment_pipeline("Miami, FL")
    # Replace with the unique long ID sequence found directly in your Google Sheet URL
    TARGET_SPREADSHEET_ID = "your_google_sheet_id_here"
    
    run_zillow_google_sheets_pipeline(TARGET_SPREADSHEET_ID, "Miami, FL")

    run_global_pipeline()