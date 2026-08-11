# AI SALES AGENT
## - Includes:
# - Generate Lead Scope - Google Maps, X, LinkedIn, Reddit Instagram, TikTok, Quora.
# - Multi-client SaaS mode
# - Lead scoring & filtering
# - LinkedIn DM generation
# - Excel export + notifications
# - Stripe subscription gating
# - NO-WEBSITE LEAD TARGETING
# - Maps Link Generation
# - Generate Tailored Outreach - AI written SMS copy, and CRM Sync Notion
# - Notion CRM sync (attach prompt, Loom script, pricing)
# - AI-generated Web App Prompt
# - AI-written Loom-style video script - Video consultation
# - AI-generated SMS copy
# - Auto proposal pricing logic

## - Overall Architecture:
# - Google Sheets (lead source + AI prompt storage)
# - Notion CRM sync (attach prompt, Loom script, pricing)
# - AI-generated Web App Prompt
# - AI-written Loom-style video script
# - AI-generated SMS copy
# - Auto proposal pricing logic

# Scrape Lead
# → Detect No Website
# → Generate Tailored Outreach
# → Generate AI Web App Prompt
# → Store in Google Sheet
# → Notify You (Email / SMS / Telegram)

# Each lead gets:

# * `source`
# * `platform`
# * `contact_type` (Email vs DM)

# FUTURE ENHANCEMENTS
# Add the following lead scopes for getting first time users and customers: Reddit Instagram, TikTok, Quora 


import os
import stripe
import datetime as dt_module
import openai
import requests
import gspread
import json
import hashlib
import csv
import urllib
import logging
from real_estate import fetch_zillow_properties
from datetime import datetime as dt
from typing import List, Dict, Any

datetime = dt_module

import gspread
from google import genai
from openai import OpenAI
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas
from google.oauth2.service_account import Credentials
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from openpyxl import load_workbook
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials, ServiceAccountCredentials
from twilio.rest import Client

from googleapiclient.discovery import build


# =========================
# CONFIG
# =========================
# ---------------- Configure logging to output to standard out for GitHub Actions ---------------- #
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger(__name__)



# ---------------- INTIALIZE CLIENT ---------------- #
openai.api_key = os.getenv("OPENAI_API_KEY")
openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY")) 
gemini_client = genai.Client() # The SDK automatically detects the GEMINI_API_KEY environment variable
stripe.api_key = os.getenv("STRIPE_SECRET_KEY")

# ---------------- ENV SETUP ---------------- #
SERPAPI_KEY = os.getenv("SERPAPI_API_KEY")
SENDGRID_KEY = os.getenv("SENDGRID_API_KEY")
NOTION_API_KEY = os.getenv("NOTION_API_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
GOOGLE_SERVICE_ACCOUNT_FILE="service_account.json"
GOOGLE_SHEET_ID = "xxxx"
GOOGLE_SHEET_NAME = "TECH_LEADS"
GOOGLE_CALENDAR_ID = "primary"
NOTION_DB_ID = os.getenv("NOTION_DB_ID")
CLOUD_SHEET_FILE = "cloud_sheet.xlsx"
STATE_FILE = "state.json" # TRACK ROW COUNT IN EXCEL STATE FILE
GC_QUOTA_LIMIT = 50

#config base case
if os.path.exists("config.json"):
	with open("config.json") as f:
		CONFIG = json.load(f)
else:
  raise FileNotFoundError("config.json not found. Please create a config.json file with the necessary configuration.")

# ---------------- DAILY GOOGLE CLOUD QUOTA TRACKER ---------------- #
# ---------------- Google API daily budget guard for once-daily GitHub Actions runs. ---------------- #
GOOGLE_DAILY_QUOTA_LIMIT = max(
    1,
    int(GC_QUOTA_LIMIT)
)


def _load_google_quota_state():
    today_str = datetime.date.today().isoformat()
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r") as f:
                state = json.load(f)
            if state.get("date") == today_str:
                return {"date": today_str, "used": int(state.get("used", 0))}
        except Exception:
            pass
    return {"date": today_str, "used": 0}


def _save_google_quota_state(state):
    state["date"] = datetime.date.today().isoformat()
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)


def reserve_google_call(operation_name, amount=1):
    state = _load_google_quota_state()
    used = int(state.get("used", 0))
    if used + amount > GOOGLE_DAILY_QUOTA_LIMIT:
        logger.warning(
            "Google daily quota exhausted (%s/%s). Skipping %s.",
            used,
            GOOGLE_DAILY_QUOTA_LIMIT,
            operation_name,
        )
        return False
    state["used"] = used + amount
    _save_google_quota_state(state)
    logger.info(
        "Reserved %s Google call(s) for %s. Remaining: %s",
        amount,
        operation_name,
        GOOGLE_DAILY_QUOTA_LIMIT - state["used"],
    )
    return True


def get_google_quota_status():
    state = _load_google_quota_state()
    used = int(state.get("used", 0))
    return {
        "used": used,
        "limit": GOOGLE_DAILY_QUOTA_LIMIT,
        "remaining": max(0, GOOGLE_DAILY_QUOTA_LIMIT - used),
    }

# Google Maps API daily budget guard.
def get_and_update_daily_count():
    """Reserve one slot from the Google daily quota for a Maps/Google API request."""
    return reserve_google_call("google_maps_request", amount=1)

# ---------------- GOOGLE AUTH ---------------- #
# Unified Google Services Authentication Matrix
try:
    creds = Credentials.from_service_account_file(
        GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/calendar",
            "https://www.googleapis.com/auth/drive"
        ]
    )
    gc = gspread.authorize(creds)
    if reserve_google_call("google_sheets_open", amount=1):
        sheet = gc.open(GOOGLE_SHEET_NAME).sheet1
    else:
        sheet = None
    calendar_service = build("calendar", "v3", credentials=creds)
    drive_service = build("drive", "v3", credentials=creds)
except Exception as e:
    print(f"⚠️ Google Service Account auth skipped or failed: {e}")
    sheet = None
    calendar_service = None
    drive_service = None

# ==========================================
# AI COPYWRITING & TRANSFORM OPERATIONS
# ==========================================
# ---------------- OPENAI MODEL CONFIG ---------------- #
def openai_generate(prompt, temperature=0.7):
	try:
		response = openai_client.chat.completions.create(
			model="gpt-4o-mini",
			messages=[{"role": "user", "content": prompt}],
			temperature=temperature
		)
		return response.choices[0].message.content.strip()
	except Exception as e:
		print(f"AI Generation error: {e}")
		return "[Generation Failure Placeholder]"


# ---------------- GEMINI MODEL CONFIG ---------------- #
def gemini_generate(prompt, temperature=0.7):
	try: 
		response = gemini_client.generate_content(
			model="gemini-2.5-flash",
			contents=prompt
		)
		return response.text
	except Exception as e:
		print(f"Gemini Generation error: {e}")
		return "[Generation Failure Placeholder]"

#--- switch between openai and gemini models ------- #
def ai_model_selection(prompt, temperature=0.7, model="gemini"):
    if model == "openai":
        return openai_generate(prompt, temperature)
    elif model == "gemini":
        return gemini_generate(prompt, temperature)
    else:
        raise ValueError(f"Invalid model: {model}")

# ---------------- LOOM SCRIPT GENERATOR ---------------- #
def build_loom_script(lead):
    return f"""
Write a casual Loom-style sales video script.

Lead name: {lead['name']}
Business: {lead['business']}
Pain point: {lead['pain']}
Offer: {lead['offer']}

Tone: friendly, confident, personalized.
Under 90 seconds.
"""

# ---------------- SMS COPY GENERATION ---------------- #
def build_sms_copy(lead):
    return f"""
Write a personalized SMS outreach message.

Recipient: {lead['name']}
Business: {lead['business']}
Pain point: {lead['pain']}

Goal: spark curiosity and reply.
Max 2 sentences.
"""

# =========================
# PRICING ENGINE
# =========================
# ---------------- PRICING ENGINE ---------------- #
def calculate_price(size, urgency, custom):
    base = {
        "solo": 2000,
        "smb": 4500,
        "enterprise": 10000
    }.get(size, 3000)

    if urgency == "high":
        base *= 1.2
    if custom == "yes":
        base *= 1.3

    return int(base)

# =========================
# CRM SYNC & LEAD MANAGEMENT
# =========================

# ---------------- CRM UTILITY ---------------- #
# EXPORT SHEET TO EXCEL
def export_sheet_to_excel(spreadsheet_id=GOOGLE_SHEET_ID, sheet_name=GOOGLE_SHEET_NAME):
    """Exports Google Sheet matrix targets locally and strictly returns the file string path."""
    today_str = datetime.date.today().isoformat()
    file_name = f"leads_{today_str}.xlsx"
    if not drive_service:
        # Fallback empty generation so local max_row validation doesn't crash program execution
        wb = load_workbook("template.xlsx") if os.path.exists("template.xlsx") else load_workbook()
        wb.save(file_name)
        return file_name

    if not reserve_google_call("google_drive_export", amount=1):
        logger.warning("Skipping Google Drive export because the daily Google quota is exhausted.")
        return file_name

    request = drive_service.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    with open(file_name, "wb") as f:
        f.write(request.execute())
    print(f"Exported Excel file: {file_name}")
    return file_name

    
    

# ---------------- NOTION SYNC ---------------- #
def push_to_notion(lead, artifacts):
    url = "https://api.notion.com/v1/pages"
    headers = {
        "Authorization": f"Bearer {NOTION_API_KEY}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json"
    }

    data = {
        "parent": {"database_id": NOTION_DB_ID},
        "properties": {
            "Name": {"title": [{"text": {"content": lead["business"]}}]},
            "Location": {"rich_text": [{"text": {"content": lead["location"]}}]},
            "Website": {"url": lead["website"]},
            "Niche": {"select": {"name": lead["niche"]}},
            "Price": {"number": artifacts["price"]},
            "Lead Type": {"select": {"name": lead['Lead Type']}},
            "profileUrl": {"url": lead['Google Maps Link']},
            "AI Prompt": {"rich_text": [{"text": {"content": artifacts["web_prompt"][:2000]}}]},
            "Loom Script": {"rich_text": [{"text": {"content": artifacts["loom"][:2000]}}]},
            "SMS Copy": {"rich_text": [{"text": {"content": artifacts["sms"]}}]},

        }
    }

    requests.post(url, headers=headers, json=data)

# ---------------- CALENDAR EVENT CREATION ---------------- #
def create_calendar_event(service, lead):
    event = {
        "summary": f"Sales Call – {lead['business']}",
        "description": "Discovery + walkthrough",
        "start": {"dateTime": "2026-02-10T14:00:00"},
        "end": {"dateTime": "2026-02-10T14:30:00"}
    }
    if not reserve_google_call("google_calendar_event_create", amount=1):
        return "https://calendar.google.com"
    event = service.events().insert(calendarId="primary", body=event).execute()
    return event.get("htmlLink")

# ---------------- AI EMAIL GENERATOR ---------------- #
def generate_email(business):
    prompt = f"""
    Write a personalized cold outreach email to:
    Business: {business['title']}
    Website: {business.get('website')}
    Location: {business.get('address')}
    Goal: Book a short sales call
    Tone: Professional, direct, friendly
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content.strip()

def generate_followup(business, step):
    prompt = f"Write follow-up #{step} for {business}. Keep it short."
    res = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )
    return res.choices[0].message.content.strip()

# =========================
# INVOICE GENERATION
# =========================
def has_active_subscription(customer_id):
    subs = stripe.Subscription.list(customer=customer_id, status="active")
    return len(subs.data) > 0

def create_stripe_checkout(lead_name, price):
    session = stripe.checkout.Session.create(
        payment_method_types=["card"],
        line_items=[{
            "price_data": {
                "currency": "usd",
                "product_data": {
                    "name": f"{lead_name} – Custom Web App Build"
                },
                "unit_amount": price * 100,
            },
            "quantity": 1,
        }],
        mode="payment",
        success_url=f"{os.getenv('PUBLIC_CHECKOUT_DOMAIN')}/success",
        cancel_url=f"{os.getenv('PUBLIC_CHECKOUT_DOMAIN')}/cancel",
    )
    return session.url
    


# =========================
# AI GENERATION
# =========================

# ---------------- AI PROMPT GENERATOR ---------------- #

def build_web_app_prompt(lead):
    return f"""
You are a senior conversion-focused product designer and full-stack engineer.

Build a high-converting web app for:
Business: {lead['business']}
Industry: {lead['industry']}
Target Customer: {lead['avatar']}
Offer: {lead['offer']}
Pain Point: {lead['pain']}
Desired Outcome: {lead['outcome']}

Use this structure:
1. Hero
2. Success State
3. Problem-Agitate-Transition
4. Value Stack
5. Social Proof
6. Transformation
7. Secondary CTA
8. Footer

Optimize for speed, clarity, and conversions.
"""

# ---------------- DEMO SITE GENERATION ---------------- #
def generate_demo_site(lead):
    prompt = f"""
Generate a single-page HTML website for:
Business: {lead['business']}
Offer: {lead['offer']}
CTA: Book a Call

Use modern Tailwind-style layout.
"""
    html = openai_generate(prompt)
    file_name = f"demo_{lead['business']}.html"
    with open(file_name, "w") as f:
        f.write(html)
    return file_name



def generate_x_dm(tweet_text, niche):
    prompt = f"""
    Write a short Twitter DM based on this tweet:

    "{tweet_text}"

    Niche: {niche['name']}
    Goal: Start a conversation, NOT pitch.
    """
    res = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )
    return res.choices[0].message.content.strip()


# =========================================
# LEAD CLASSIFICATION & MESSAGE GENERATION
# =========================================

# ---------------- NO WEBSITE MESSAGE GENERATION ---------------- #
def generate_no_website_message(lead):
    return f"""Hi {lead.get('name','there')},

I noticed your business in {lead.get('location')} doesn’t currently have a dedicated website.

We help businesses like yours attract more customers from Google, build trust online, and automate inquiries.

Would you be open to a quick walkthrough showing what this could look like for {lead.get('business_name')}?

Best,
{{Your Name}}
"""
def generate_ai_no_website_message(lead):
    prompt = f"""Write a concise, professional Email to a business owner.
Business: {lead['Business Name']}
Location: {lead['Location']}
Context: Business does not have a website.
Tone: Helpful, non-salesy."""
    return ai_model_selection(prompt, model="openai")

def generate_ai_no_website_low_review(lead):
    prompt = f"""Write a concise, professional Email to a business owner.
Business: {lead['Business Name']}
Location: {lead['Location']}
Context: Business does not have a website and has low reviews.
Tone: Helpful, non-salesy."""
    return ai_model_selection(prompt, model="openai")

def generate_ai_website_low_review(lead):
    prompt = f"""Write a concise, professional Email to a business owner.
Business: {lead['Business Name']}
Location: {lead['Location']}
Context: Business has a website but has low reviews.
Tone: Helpful, non-salesy."""
    return ai_model_selection(prompt, model="openai")


# ---------------- WEBSITE CHECK ---------------- #
def has_website(lead):
    website = lead.get("website")
    if not website:
        return False
    if any(s in website.lower() for s in ["facebook.com", "instagram.com", "linkedin.com"]):
        return False
    return True

# ---------------- LEAD CLASSIFICATION ---------------- #
def classify_lead(lead):
    if not has_website(lead):
        return "NO_WEBSITE"
    elif not has_website(lead) and lead.get("reviews", 0) < 10:
        return "NO_WEBSITE_LOW_REVIEW"
    elif not has_website(lead) and lead.get("reviews", 0) < 10 and lead.get("rating", 0) <= 3:
        return "NO_WEBSITE_LOW_REVIEW_LOW_RATING"
    elif not has_website(lead) and lead.get("reviews", 0) >= 15 and lead.get("rating", 0) >= 4:
        return "NO_WEBSITE_HIGH_REVIEW_HIGH_RATING"
    elif has_website(lead) and lead.get("reviews", 0) < 10:
        return "HAS_WEBSITE_LOW_REVIEW"
    elif has_website(lead) and (lead.get("reviews", 0) < 10 and lead.get("rating", 0) <= 3):
        return "HAS_WEBSITE_LOW_REVIEW_LOW_RATING"
    elif has_website(lead) and lead.get("reviews", 0) >= 15 and lead.get("rating", 0) >= 4:
        return "HAS_WEBSITE_HIGH_REVIEW_HIGH_RATING"
    elif has_website(lead):
        return "HAS_WEBSITE"
    else:
        return "NO LEAD CLASSIFICATION"


def process_lead(lead):
    lead["lead_type"] = classify_lead(lead)
    if lead["lead_type"] == "NO_WEBSITE" or lead["lead_type"] == "NO_WEBSITE_HIGH_REVIEW_HIGH_RATING":
        lead["tailored_message"] = generate_ai_no_website_message(lead)
    elif lead["lead_type"] == "NO_WEBSITE_LOW_REVIEW" or lead["lead_type"] == "NO_WEBSITE_LOW_REVIEW_LOW_RATING":
        lead["tailored_message"] = generate_ai_no_website_low_review(lead)
    elif lead["lead_type"] == "HAS_WEBSITE_LOW_REVIEW" or lead["lead_type"] == "HAS_WEBSITE_LOW_REVIEW_LOW_RATING":
        lead["tailored_message"] = generate_ai_website_low_review(lead)
    return lead

def score_lead(lead):
    score = 0
    if lead.get("lead_type") == "NO_WEBSITE":
        score += 40
    if lead.get("rating", 0) >= 4:
        score += 20
    if lead.get("reviews", 0) >= 20:
        score += 20
    return score

# def has_website(lead, query, lead_source=None):
#     if lead_source == "google_maps" and lead != None:
#         if lead:
#             website = lead.get("website")
#             if not website:
#                 classify_lead(False) #return False
#             if any(s in website.lower() for s in ["facebook.com", "instagram.com", "linkedin.com"]):
#                 classify_lead(False) #return False
#             return classify_lead(False) #return True
#         elif (lead != True or lead == None) and query and GOOGLE_MAPS_API_KEY:
            
#     elif lead_source == "linkedin":
#         website = lead.get("website")
#         if not website:
#             return classify_lead(False)
#         if any(s in website.lower() for s in ["facebook.com", "instagram.com", "linkedin.com"]):
#             return classify_lead(False)
#         return classify_lead(True)
#     elif lead_source == "x":
#         return classify_lead(True)
#     else:
#         print(ValueError("Invalid lead source specified for website check."))
#         return None


# ---------------- AI MESSAGE GENERATOR ---------------- #
def generate_ai_sms_message(lead):
    prompt = f"""Write a concise, friendly SMS to a business owner.
Business: {lead['Business Name']}
Location: {lead['Location']}
Context: Business does not have a website.
Tone: Helpful, non-salesy."""
    return ai_model_selection(prompt, model="openai")

# ---------------- LINKEDIN DM GENERATOR ---------------- #
def generate_linkedin_dm(lead, niche, location):
    prompt = f"""
    Write a short, natural LinkedIn DM.

    Recipient:
    - Name: {lead.get('name')}
    - Role: {lead.get('title')}
    - Company: {lead.get('company')}
    - Location: {location}

    Niche: {niche['name']}
    Value Proposition: {niche['value_prop']}

    Rules:
    - Max 3 short sentences
    - NO selling
    - NO links
    - Sound human, casual, respectful
    - Goal: Start a conversation, not book a call

    End with a soft question.
    """

    # response = openai.ChatCompletion.create(
    #     model="gpt-4",
    #     messages=[{"role": "user", "content": prompt}],
    #     temperature=0.7
    # )

    # return response.choices[0].message.content.strip()
    ai_model_selection(prompt, model="openai")

# =========================================
# LEAD SCOPE GENERATION
# =========================================
# ---------------- FETCH METRICS FROM OUTREACH SOURCE ---------------- #
def fetch_metrics_from_source(source_type, client, kwargs):
    """
    Helper function to dynamically pull metrics based on the target platform.
    """
    # Fallback default values
    total_count, new_count = "N/A", "N/A"
    
    try:
        if source_type == "google_sheets":
            if not reserve_google_call("google_sheets_metrics_read", amount=1):
                logger.warning("Skipping Google Sheets metrics fetch because the daily Google quota is exhausted.")
                return "N/A", "N/A"
            # Expecting client to be an authorized gspread client
            sheet_name = kwargs.get("sheet_name", "LeadsSheet")
            sheet = client.open(sheet_name).sheet1
            total_count = len(sheet.get_all_values()) - 1
            new_count = "Dynamic Fetch"  # Replace with explicit cell fetch if tracked
            
        elif source_type == "notion":
            # Expecting client to be a notion_client.Client instance
            database_id = kwargs.get("database_id")
            response = client.databases.query(database_id=database_id)
            total_count = len(response.get("results", []))
            new_count = "Dynamic Fetch"  # Replace with filtering logic for 'New' status
            
        elif source_type == "outreach":
            # Expecting client to be an Outreach API client instance
            # Example logic leveraging an Outreach SDK wrapper
            response = client.prospects.list()
            total_count = response.get("meta", {}).get("count", "N/A")
            new_count = "Dynamic Fetch"
            
    except Exception as e:
        logger.error(f"⚠️ Failed to fetch metrics from {source_type}: {e}")
        
    return total_count, new_count

# ---------------- GOOGLE MAPS SCRAPER ---------------- #
def scrape_google_maps(api: str, query: str, location: str, limit: int = 10) -> List[Dict[str, Any]]:
    """Fetches local business data from SerpAPI Maps engine, normalizing keys."""
    params = {"engine": "google_maps", "q": query, "location": location, "api_key": SERPAPI_KEY}
    try:
        if api == "serpapi":
            if not get_and_update_daily_count():
                return None
            logger.info(f"Initiating Google Places API search for query: '{query}'")
            try:
                res = requests.get("https://serpapi.com/search", params=params, timeout=15).json()
                raw_results = res.get("local_results", [])[:limit]
                normalized = []
                for item in raw_results:
                    normalized.append({
                        "company": item.get("title", "Unknown Local Business"),
                        "website": item.get("website", ""),
                        "phone": item.get("phone", ""),
                        "profileUrl": item.get("gps_coordinates", {}).get("links", "N/A"),
                        "industry": item.get("type", "Local Business"),
                        "pain": "No prominent digital presence matching search metrics",
                    })
                return normalized
            except requests.exceptions.RequestException as e:
                logger.error(f"SerpAPI Network Error: {e}")
                print("::endgroup::")
                return None
        elif api == "googleapi":
            if not get_and_update_daily_count():
                return None
            params = {"engine": "google", "q": query, "api_key": GOOGLE_MAPS_API_KEY}
            url = "https://googleapis.com"
            headers = {
                "Content-Type": "application/json",
                "X-Goog-Api-Key": GOOGLE_MAPS_API_KEY,
                "X-Goog-FieldMask": "places.displayName,places.websiteUri"
            }
            payload = {
                "textQuery": query,
                "maxResultCount": limit
            }
            logger.info(f"Initiating Google Places API search for query: '{query}'")
            try:
                response = requests.post(url, json=payload, headers=headers)
                response.raise_for_status()
                #extract all the results similar to the serpapi logic but using the google maps api response structure and then check if they have a website listed or not and return true or false accordingly
                results = res.get("places", [])
                logger.info(f"Successfully retrieved {len(raw_results)} raw results.")
                normalized = []
                for item in results:
                    normalized.append({
                        "company": item.get("displayName", {}).get("text", "Unknown Local Business"),
                        "website": item.get("websiteUri", ""),
                        "phone": item.get("nationalPhoneNumber", ""),
                        "profileUrl": item.get("googleMapsUri", "N/A"),
                        "industry": item.get("primaryType", "Local Business"),
                        "pain": "No prominent digital presence matching search metrics",
                    })
                logger.info("Data extraction and normalization complete.")
                return normalized     
            except requests.exceptions.RequestException as e:
                logger.error(f"API Request failed: {e}")
                print("::endgroup::")
                return None
        else:
            raise ValueError("Invalid API specified for Google Maps scraping. Use 'serpapi' or 'googleapi'.")
    except Exception as e:
        print(f"Google Maps scrape exception: {e}")
        return []

# ---------------- LINKEDIN LEAD FETCHER ---------------- #
def fetch_linkedin_leads(phantom_id: str, api_key: str) -> List[Dict[str, Any]]:
    url = f"https://api.phantombuster.com/api/v2/agents/fetch-output?id={phantom_id}"
    headers = {"X-Phantombuster-Key": api_key}
    logger.info(f"Initiating LinkedIn lead fetch for ID: {phantom_id}")
    try:
        res = requests.get(url, headers=headers, timeout=15).json()
        raw_data = res.get("data", [])
        normalized = []
        for item in raw_data:
            normalized.append({
                "company": item.get("company", "Unknown Organization"),
                "website": item.get("website", ""),
                "phone": item.get("phone", ""),
                "email": item.get("email", ""),
                "profileUrl": item.get("profileUrl", "https://linkedin.com"),
                "industry": item.get("industry", "Corporate Enterprise"),
                "pain": "Optimizing B2B outreach conversion channels"
            })

        return normalized
    except Exception as e:
        logger.error(f"LinkedIn fetch exception: {e}")
        print("::endgroup::")
        return []

# ---------------- X LEAD SCRAPER ---------------- #
def scrape_x_leads(query: str) -> List[Dict[str, Any]]:
    headers = {"Authorization": f"Bearer {os.getenv('X_BEARER_TOKEN')}"}
    params = {"query": query, "max_results": 10, "tweet.fields": "author_id,id"}
    logger.info(f"Initiating X (Twitter) scrape for query: '{query}'")
    try:
        res = requests.get("https://api.twitter.com/2/tweets/search/recent", headers=headers, params=params, timeout=15).json()
        raw_data = res.get("data", [])
        normalized = []
        for item in raw_data:
            normalized.append({
                "company": f"X User Handle: {item.get('author_id')}",
                "website": "",
                "phone": "",
                "email": f"{item.get('author_id')}@fallback-x-lead.com",  # Mock formatting for validation
                "profileUrl": f"https://x.com/{item.get('author_id')}/status/{item.get('id')}",
                "industry": "Digital/Social Influence Channel",
                "pain": "Converting real-time conversational social intent"
            })
        return normalized
    except Exception as e:
        logger.error(f"X (Twitter) scrape exception: {e}")
        print("::endgroup::")
        return []


# =========================
# SEND NOTIFICATIONS
# =========================

# ---------------- EMAIL NOTIFICATION ---------------- #
# def notify_email(new_count, total_count):
#     subject = "🚀 New Lead Added"
#     body = f"""
#     A new lead has been added to your outreach list.

#     New entries: {new_count}
#     Total leads: {total_count}

#     Check your Excel file for details.
#     """

#     message = Mail(
#         from_email=os.getenv("NOTIFY_EMAIL_FROM"),
#         to_emails=os.getenv("NOTIFY_EMAIL_TO"),
#         subject=subject,
#         plain_text_content=body
#     )
#     sg = SendGridAPIClient(os.getenv("SENDGRID_API_KEY"))
#     sg.send(message)
#     print(f"📧 Sending SendGrid Digest Alert: {new_count} newly injected platform leads added.")


def notify_email(lead, source_type="google_sheets", db_client=None, **kwargs):
    """
    Sends an email notification for a new lead, dynamically fetching 
    metrics based on the active source pipeline (Google Sheets, Notion, or Outreach).
    """
    # body = f"""
    # A new lead has been added to your outreach list.

    # New entries: {new_count}
    # Total leads: {total_count}

    # Check your Excel file for details.
    # """
    
    # 1. Dynamically pull metrics depending on the platform source
    total_count, new_count = fetch_metrics_from_source(source_type, db_client, kwargs)

    # 2. Map source type to a clean display name for the email template
    source_names = {
        "google_sheets": "Google Sheets file",
        "notion": "Notion CRM Board",
        "outreach": "Outreach Database"
    }
    display_source = source_names.get(source_type, "Outreach list")
    message = Mail(
        from_email=os.getenv("NOTIFY_EMAIL_FROM"),
        to_emails=os.getenv("NOTIFY_EMAIL_TO"),
        subject=f"🚀 New Lead Added to {source_type.replace('_', ' ').title()}: {lead['Business Name']}",
        # plain_text_content=body,
        html_content=f"""
        <h1>A new lead has been added to your {display_source} outreach list.</h1>
        <br/>
        <strong>Business:</strong> {lead['Business Name']}<br>
        <strong>Location:</strong> {lead['Location']}<br>
        <strong>Lead Type:</strong> {lead['Lead Type']}<br>
        <a href="{lead['Google Maps Link']}">View on Google Maps</a>
        <br/>
        <strong>New entrY:</strong> {new_count}<br>
        <strong>Total leads:</strong> {total_count}<br><br>
        <p>Check your {display_source} for details.</p>
        """
    )
    try:
        sg = SendGridAPIClient(os.getenv("SENDGRID_API_KEY"))
        sg.send(message)
        print(f"📧 Sending SendGrid Digest Alert [{source_type.upper()}]: {new_count} updated leads for {lead['Business Name']} from {lead['Location']}") # print(f"📧 Sending SendGrid Digest Alert [{source_type.upper()}]: Updated metrics for {lead['Business Name']}")
    except Exception as e:
        logger.error(f"❌ Failed to send SendGrid email: {e}")
        print("::endgroup::")
    
# ---------------- TELEGRAM NOTIFICATION ---------------- #
def notify_telegram(lead, source_type="google_sheets", db_client=None, **kwargs):
    """
    Sends a Telegram markdown notification for a new lead, dynamically fetching 
    metrics based on the active source pipeline (Google Sheets, Notion, or Outreach).
    """
    
    # 1. Dynamically pull metrics using the same core logic
    total_count, new_count = fetch_metrics_from_source(source_type, db_client, kwargs)
    message = (
        f"🚀 *New Lead Added!*\n\n"
        f"➕ New entries: {new_count}\n"
        f"📊 Total leads: {total_count}\n"
        f"*Business:* {lead['Business Name']}\n"
        f"*Location:* {lead['Location']}\n"
        f"*Type:* {lead['Lead Type']}\n"
        f"[Open in Google Maps]({lead['Google Maps Link']})"
    )

    url = f"https://api.telegram.org/bot{os.getenv('TELEGRAM_BOT_TOKEN')}/sendMessage"
    payload = {
        "chat_id": os.getenv("TELEGRAM_CHAT_ID"),
        "text": message,
        "parse_mode": "Markdown"
    }

    try:
        requests.post(url, json=payload)
        print(f"Dispatched Telegram Notification Payload [{source_type.upper()}]. Total repository count: {total_count}")
    except Exception as e:
        logger.error(f"❌ Failed to send Telegram notification: {e}")
        print("::endgroup::")


def notify_sms(lead):
    body = generate_ai_sms_message(lead)
    client = Client(
        os.getenv("TWILIO_ACCOUNT_SID"),
        os.getenv("TWILIO_AUTH_TOKEN")
    )
    client.messages.create(
        body=body,
        from_=os.getenv("TWILIO_PHONE_NUMBER"),
        to=os.getenv("ALERT_PHONE_NUMBER")
    )
    
def notify_all(lead):
    notify_email(
        lead, 
        source_type="google_sheets", 
        db_client=gspread_client, 
        sheet_name="Client Outreach CRM"
    )
    notify_telegram(
        lead, 
        source_type="google_sheets", 
        db_client=gspread_client, 
        sheet_name="Client Outreach CRM"
    )
    notify_sms(lead)



# ==========================================
# STORAGE UTILITIES & UTILS LAYER
# ==========================================
def lead_hash(email):
    return hashlib.md5(email.encode()).hexdigest()

def get_google_sheet_row_count():
    """Returns the total number of populated rows, excluding the header."""
    try:
        if not sheet or not reserve_google_call("google_sheets_read_row_count", amount=1):
            return 0
        # get_all_values() returns a list of lists representing the populated grid
        all_rows = sheet.get_all_values()

        if not all_rows:
            return 0

        # Total populated rows minus 1 for the header row
        return len(all_rows) - 1
    except Exception as e:
        print(f"Error reading Google Sheet row count: {e}")
        return 0

def already_queued(email):
    if not sheet or not reserve_google_call("google_sheets_read_dedup", amount=1):
        return False
    try:
        records = sheet.get_all_records()
        hashes = [lead_hash(r["Email"]) for r in records if r.get("Email")]
        return lead_hash(email) in hashes
    except Exception:
        return False

# ---------------- EXCEL NEW-LEAD DETECTOR ---------------- #
def load_last_row_count():
    if not os.path.exists(STATE_FILE):
        return 0
    with open(STATE_FILE, "r") as f:
        return json.load(f).get("sheet_row_count", 0)

def save_row_count(count):
    today_str = datetime.today().strftime('%Y-%m-%d')
    with open(STATE_FILE, "w") as f:
        json.dump({"date": today_str, "sheet_row_count": count}, f)

def get_excel_row_count(file_name):
	try:
		wb = load_workbook(file_name)
		return wb.active.max_row - 1 
	except Exception:
		return 0
    # return ws.max_row - 1  # exclude header
    
# ---------------- GOOGLE SHEETS NEW-LEAD DETECTOR ---------------- #
def load_state():
    if os.path.exists(STATE_FILE):
        return json.load(open(STATE_FILE))
    return {"last_row": 1}

def save_state(row):
    today_str = datetime.today().strftime('%Y-%m-%d')
    json.dump({"date": today_str, "last_row": row}, open(STATE_FILE, "w"))

def get_new_leads(sheet):
    state = load_state()
    last_row = state["last_row"]

    all_rows = sheet.get_all_records()
    current_row_count = len(all_rows) + 1  # header row

    if current_row_count <= last_row:
        return []

    new_leads = all_rows[last_row - 1 :]
    save_state(current_row_count)

    return new_leads
def process_new_sheet_entries(sheet):
    new_leads = get_new_leads(sheet)

    for lead in new_leads:
        notify_email(lead)
        notify_telegram(lead)

# ---------------- GOOGLE MAPS LINK GENERATOR ---------------- #
#example: https://www.google.com/maps/search/?api=1&query=Elite+Auto+Detailing+Dallas+TX
def generate_maps_link(business_name, location):
    query = f"{business_name} {location}"
    encoded = urllib.parse.quote_plus(query)
    return f"https://www.google.com/maps/search/?api=1&query={encoded}"


# ---------------- SEND EMAIL ---------------- #
def send_email(to_email, subject, content):
    message = Mail(
        from_email="you@yourdomain.com",
        to_emails=to_email,
        subject=subject,
        html_content=content
    )
    sg = SendGridAPIClient(SENDGRID_KEY)
    sg.send(message)

# ---------------- CREATE CALENDAR EVENT ---------------- #
def book_call(business_name: str, email: str) -> str:
    if not calendar_service:
        return "https://cal.com/fallback-booking"
    try:
        if not reserve_google_call("google_calendar_event_create", amount=1):
            return "https://calendar.google.com"
        start_time = (dt.now(datetime.timezone.utc) + datetime.timedelta(days=2)).isoformat() + "Z"
        end_time = (dt.now(datetime.timezone.utc) + datetime.timedelta(days=2, minutes=30)).isoformat() + "Z"
        event = {
            "summary": f"Intro Call - {business_name}",
            "start": {"dateTime": start_time},
            "end": {"dateTime": end_time},
            "attendees": [{"email": email}]
        }
        res = calendar_service.events().insert(calendarId=GOOGLE_CALENDAR_ID, body=event).execute()
        return res.get("htmlLink", "https://calendar.google.com")
    except Exception:
        return "https://calendar.google.com"

# ---------------- FOLLOW-UP GENERATOR ---------------- #
def generate_follow_up(business_name, step):
    prompt = f"""
    Write follow-up email #{step} for {business_name}.
    Keep it short and polite.
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content.strip()

def generate_proposal_pdf(lead, price):
    file_name = f"proposal_{lead['business']}.pdf"
    c = canvas.Canvas(file_name, pagesize=LETTER)

    text = c.beginText(40, 750)
    text.textLine(f"Proposal for {lead['business']}")
    text.textLine("")
    text.textLine(f"Problem: {lead['pain']}")
    text.textLine(f"Solution: {lead['offer']}")
    text.textLine(f"Investment: ${price}")
    text.textLine("Timeline: 14–21 days")
    text.textLine("")
    text.textLine("Let’s build something powerful.")

    c.drawText(text)
    c.save()
    return file_name
    

# =========================
# MAIN PIPELINE
# =========================

# ---------------- MAIN AGENT ---------------- #
def run_agent():
    today = str(datetime.date.today())

    sources = ["google_maps",
            #    "linkedin", 
            #    "x"
               ]

    for source in sources:
        for niche in CONFIG.get("tech config", {}).get("niches", []):
            for location in CONFIG.get("locations", []):
                # Fetch routing phase
                if source == "google_maps":
                    leads = scrape_google_maps("googleapis",niche["search_query"], location, CONFIG.get("daily_limit_per_combo", 10))
                    source_label = "Google Maps"
                elif source == "linkedin":
                    leads = fetch_linkedin_leads(os.getenv("PHANTOM_ID", ""), os.getenv("PHANTOMBUSTER_API_KEY", ""))
                    source_label = "LinkedIn"
                elif source == "x":
                    query = niche.get("x_query", f"{niche['search_query']} {location}")
                    leads = scrape_x_leads(query)
                    source_label = "X"

                # Transformation matrix processing phase
                for lead in leads:
                    # Clean the lead key-structure slightly to match your layout checks
                    # Ensure location and business names carry safely into classification frameworks
                    lead["Location"] = location
                    lead["Company"] = lead.get("company", "Unknown Business")

                    # --- Step 1: Execute Website Check, Classification & Scoring ---
                    lead = process_lead(lead)
                    lead_score = score_lead(lead)
                    lead_type = lead.get("lead_type", "WEBSITE_CHECK_FAILED")
            
                    # Deduplication filtering logic               
                    email = lead.get("email") or f"info@{lead.get('Company').lower().replace(' ','')}.com"
                    if already_queued(email):
                        continue

                    # High level copy synthesis generation routines
                    # initial = generate_email(lead["Company"], niche, location)
                    # --- Step 2: High level copy synthesis generation routines ---
                    # Use the specific tailored message if it was generated by the classifier
                    if lead_type == "NO_WEBSITE" and "tailored_message" in lead:
                        initial = lead["tailored_message"]
                    else:
                        initial = generate_email(lead["Company"], niche, location)
                    follow1 = generate_followup(lead["Company"], 1)
                    follow2 = generate_followup(lead["Company"], 2)
                    
                    web_prompt = ai_model_selection(build_web_app_prompt(lead))
                    loom_script = ai_model_selection(build_loom_script(lead))
                    sms_copy = ai_model_selection(build_sms_copy(lead))
                    calendar_link = book_call(lead["Company"], email)

                    row_payload = [
                        niche["name"],
                        location,
                        lead["company"],
                        lead.get("website", ""),
                        lead.get("phone", ""),
                        email,
                        initial,
                        follow1,
                        follow2,
                        calendar_link,
                        "Queued",
                        today,
                        source_label,
                        lead.get("profileUrl", "N/A"),
                        web_prompt,
                        loom_script,
                        sms_copy,
                        lead_type,     # Added classification column data
                        lead_score     # Added prioritization scoring data
                    ]

                    if sheet:
                        if reserve_google_call("google_sheets_append_row", amount=1):
                            sheet.append_row(row_payload)
                            print(f"✅  Lead: {lead["company"]} - {lead["location"]} - {niche["name"]} | Score: {lead_score}] lead row: {lead['Company']} from {source_label} that {lead_type}") 
                        else:
                            logger.warning("Skipping Google Sheets append for %s because the daily Google quota is exhausted.", lead["Company"])



if __name__ == "__main__":
	
    # 1. Save the row count BEFORE running the morning scrape
    previous_count = load_last_row_count()
    # Explicit convert Google Sheets To Excel extraction configuration
    # excel_file = export_sheet_to_excel(SPREADSHEET_ID)
    # current_count = get_excel_row_count(excel_file)

    # 2. Run your scraper agent to append new data rows        
    run_agent()
    # 3. Get the updated row count directly from Google Sheets
    current_count = get_google_sheet_row_count()
                
    # 4. Compare and fire notifications if new leads exist
    if current_count > previous_count:
        new_entries = current_count - previous_count
        notify_email(new_entries, current_count)
        notify_telegram(new_entries, current_count)

    # 5. Persist the current count to your state.json file
    save_row_count(current_count)

# =========================
# MAIN PIPELINE
# =========================


# import time
# import random
# from googleapiclient.errors import HttpError

# # --- GUARDRAILS CONFIGURATION ---
# # Sheets/Drive have a default limit of 60 requests per minute per user.
# # We set a safe buffer (e.g., 45 requests per minute -> 1 request every 1.33 seconds)
# SAFE_DELAY_SECONDS = 1.4  
# MAX_RETRIES = 5

# def rate_limited_execute(request):
#     """
#     Executes a Google API request with built-in rate limiting 
#     and exponential backoff for quota errors (HTTP 429).
#     """
#     # Guardrail 1: Enforce a fixed delay between sequential requests
#     time.sleep(SAFE_DELAY_SECONDS)
    
#     for attempt in range(MAX_RETRIES):
#         try:
#             return request.execute()
            
#         except HttpError as error:
#             # Check if error is due to Rate Limiting (429) or Server Errors (5xx)
#             if error.resp.status in:
#                 # Guardrail 2: Exponential backoff with jitter
#                 sleep_time = (2 ** attempt) + random.uniform(0, 1)
#                 print(f"⚠️ Quota hit or server busy. Retrying in {sleep_time:.2f} seconds...")
#                 time.sleep(sleep_time)
#             else:
#                 # Raise other HTTP errors immediately (e.g., 403 Forbidden, 404 Not Found)
#                 raise error
                
#     raise Exception("❌ Request failed after maximum retries due to quota exhaustion.")

# # --- APPLICATION EXAMPLES ---

# def update_spreadsheet_batched(service, spreadsheet_id, range_name, values):
#     """
#     Guardrail 3: Batch data instead of updating cell-by-cell.
#     """
#     body = {
#         'values': values
#     }
    
#     # This prepares the request but does NOT execute it yet
#     request = service.spreadsheets().values().update(
#         spreadsheetId=spreadsheet_id, 
#         range=range_name, 
#         valueInputOption="RAW", 
#         body=body
#     )
    
#     # Execute safely through our guardrail function
#     return rate_limited_execute(request)

#Previous Implementation

# def run_agent():
#     today = str(datetime.date.today())

#     sources = ["google_maps", "linkedin", "x"]

#     for source in sources:
#         for niche in CONFIG.get("niches", []):
#             for location in CONFIG.get("locations", []):
                
#                 # Fetch routing phase
#                 if source == "google_maps":
#                     leads = scrape_google_maps(niche["search_query"], location, CONFIG.get("daily_limit_per_combo", 10))
#                     source_label = "Google Maps"
#                 elif source == "linkedin":
#                     leads = fetch_linkedin_leads(os.getenv("PHANTOM_ID", ""), os.getenv("PHANTOMBUSTER_API_KEY", ""))
#                     source_label = "LinkedIn"
#                 elif source == "x":
#                     query = niche.get("x_query", f"{niche['search_query']} {location}")
#                     leads = scrape_x_leads(query)
#                     source_label = "X"

#                 # Transformation matrix processing phase
#                 for lead in leads:
#                     # Clean the lead key-structure slightly to match your layout checks
#                     # Ensure location and business names carry safely into classification frameworks
#                     lead["Location"] = location
#                     lead["Business Name"] = lead.get("company", "Unknown Business")

#                     # --- Step 1: Execute Website Check, Classification & Scoring ---
#                     lead = process_lead(lead)
#                     lead_score = score_lead(lead)
#                     lead_type = lead.get("lead_type", "HAS_WEBSITE")

#                     # Deduplication filtering logic 
#                     email = lead.get("email") or f"info@{lead['Business Name'].lower().replace(' ','')}.com"
#                     if already_queued(email):
#                         continue

#                     # --- Step 2: High level copy synthesis generation routines ---
#                     # Use the specific tailored message if it was generated by the classifier
#                     if lead_type == "NO_WEBSITE" and "tailored_message" in lead:
#                         initial = lead["tailored_message"]
#                     else:
#                         initial = generate_email(lead["Business Name"], niche, location)

#                     follow1 = generate_followup(lead["Business Name"], 1)
#                     follow2 = generate_followup(lead["Business Name"], 2)
                    
#                     web_prompt = ai_generate(build_web_app_prompt(lead))
#                     loom_script = ai_generate(build_loom_script(lead))
#                     sms_copy = ai_generate(build_sms_copy(lead))
#                     calendar_link = book_call(lead["Business Name"], email)

#                     # --- Step 3: Append Processed Payload Matrix ---
#                     row_payload = [
#                         niche["name"],
#                         location,
#                         lead["Business Name"],
#                         lead.get("website", ""),
#                         lead.get("phone", ""),
#                         email,
#                         initial,
#                         follow1,
#                         follow2,
#                         calendar_link,
#                         "Queued",
#                         today,
#                         source_label,
#                         lead.get("profileUrl", "N/A"),
#                         web_prompt,
#                         loom_script,
#                         sms_copy,
#                         lead_type,     # Added classification column data
#                         lead_score     # Added prioritization scoring data
#                     ]

#                     if sheet:
#                         sheet.append_row(row_payload)
#                         print(f"✅ Injected classified [{lead_type} | Score: {lead_score}] lead row: {lead['Business Name']} from {source_label}")

# ---------------- EMAIL NOTIFICATION ---------------- #
# def notify_telegram_lead(lead):
#     token = os.getenv("TELEGRAM_BOT_TOKEN")
#     chat_id = os.getenv("TELEGRAM_CHAT_ID")

#     text = (
#         f"🚀 *New Lead*\n\n"
#         f"*Business:* {lead['Business Name']}\n"
#         f"*Location:* {lead['Location']}\n"
#         f"*Type:* {lead['Lead Type']}\n"
#         f"[Open in Google Maps]({lead['Google Maps Link']})"
#     )

#     requests.post(
#         f"https://api.telegram.org/bot{token}/sendMessage",
#         json={
#             "chat_id": chat_id,
#             "text": text,
#             "parse_mode": "Markdown"
#         }
#     )
# def notify_email(new_count, total_count):
#     subject = "🚀 New Lead Added"
#     body = f"""
#     A new lead has been added to your outreach list.

#     New entries: {new_count}
#     Total leads: {total_count}

#     Check your Excel file for details.
#     """

#     message = Mail(
#         from_email=os.getenv("NOTIFY_EMAIL_FROM"),
#         to_emails=os.getenv("NOTIFY_EMAIL_TO"),
#         subject=subject,
#         plain_text_content=body
#     )
#     sg = SendGridAPIClient(os.getenv("SENDGRID_API_KEY"))
#     sg.send(message)
#     print(f"📧 Sending SendGrid Digest Alert: {new_count} newly injected platform leads added.")
# ---------------- MAIN AGENT ---------------- #
# for source in sources:
	# 	if source == "google_maps":
	# 		# 10 niches × 20 cities × 10 leads/day = 2,000 new leads/day
	# 		for niche in CONFIG["niches"]:
	# 			for location in CONFIG["locations"]:
	# 				# leads = scrape_google_maps("Marketing Agency", "New York")
	# 				leads = scrape_google_maps(
	# 					niche["search_query"],
	# 					location,
	# 					CONFIG["daily_limit_per_combo"]
	# 				)

	# 				for lead in leads:
	# 					email = lead.get("email")
	# 					source = "Google Maps"
	# 					if not email or already_queued(email):
	# 						continue
			
	# 					initial = generate_email(lead, niche, location)
	# 					follow1 = generate_followup(lead["business"], 1)
	# 					follow2 = generate_followup(lead["business"], 2)
	# 					web_prompt = ai_generate(build_web_app_prompt(lead))
	# 					loom_script = ai_generate(build_loom_script(lead))
	# 					sms_copy = ai_generate(build_sms_copy(lead))
            

	# 					calendar_link = book_call(lead["business"], email)

	# 					# send_email(email, "Quick question", initial)

	# 					sheet.append_row([
	# 						niche["name"],
	# 						location,
	# 						lead["title"], #company
	# 						lead.get("website"),
	# 						lead.get("phone"),
	# 						email,
	# 						initial,
	# 						follow1,
	# 						follow2,
	# 						calendar_link, #" ",
	# 						"Queued",
	# 						today,
	# 						source,
	# 						lead["profileUrl"],
	# 						web_prompt,
	# 						loom_script,
  #             sms_copy
	# 					])

	# 	elif source == "linkedin":
  #     # 10 niches × 20 cities × 10 leads/day = 2,000 new leads/day
	# 		for niche in CONFIG["niches"]:
	# 			for location in CONFIG["locations"]:
						
	# 				linkedin_leads = fetch_linkedin_leads(
	# 					phantom_id=os.getenv("PHANTOM_ID"),
	# 					api_key=os.getenv("PHANTOMBUSTER_API_KEY")
	# 				)
	# 				for lead in linkedin_leads:
	# 					source = "LinkedIn"
	# 					email = lead.get("email")
	# 					if not email or already_queued(email):
	# 							continue
	# 					initial = generate_email(lead, niche, location)
	# 					follow1 = generate_followup(lead["company"], 1)
	# 					follow2 = generate_followup(lead["company"], 2)
	# 					web_prompt = ai_generate(build_web_app_prompt(lead))
	# 					loom_script = ai_generate(build_loom_script(lead))
	# 					sms_copy = ai_generate(build_sms_copy(lead))

	# 					calendar_link = book_call(lead["company"], email)

	# 					# send_email(email, "Quick question", initial)
                              
	# 					sheet.append_row([
	# 						niche["name"],
	# 						lead["location"],        #lead.get("location")
	# 						lead["company"],         #lead.get("company")
	# 						lead.get("website", ""), #website
	# 						email,       #email address
	# 						lead.get("phone", ""),   #phone number
	# 						email,                   #initial email - leag.get("email")
	# 						follow1,                 #follow-up 1
	# 						follow2,                 #follow-up 2
	# 						calendar_link,           #calendar link
	# 						"Queued",                #status
	# 						today,                   #last contacted
	# 						source,                  #lead source
	# 						lead["profileUrl"],      #lead.get("profileUrl")
  #             web_prompt,
  #             loom_script,
  #             sms_copy
	# 						# lead["name"],          #lead.get("name")
	# 						# "DM",                  #outreach type
	# 						# "",
	# 						# linkedin_dm,
	# 						# "Not Sent"
	# 					])
	# 	elif source == "x":
	# 		for niche in CONFIG["niches"]:
	# 			for location in CONFIG["locations"]:
	# 				x_leads = scrape_x_leads(niche.get(f"{niche['search_query']} {location}"))
	# 				for lead in x_leads:
	# 					source_label = "X"
	# 					profile_url = f"https://x.com/{lead['author_id']}/status/{lead['id']}"
						
	# 					# Generate dynamic outreach copy tailored to the X lead context
	# 					# matching the structure used in your LinkedIn/Maps logic
	# 					# web_prompt = f"Analyze X profile for {niche['name']} in {location}"
	# 					# loom_script = f"Hey, saw your post regarding {query}... Here is how we can help with {niche['value_prop']}."
	# 					# sms_copy = f"Hi, noticed your tweet! Are you looking to scale your business in {location}?"
						
	# 					for lead in x_leads:
	# 						email = lead.get("email")
	# 						if not email or already_queued(email):
	# 								continue
									
	# 						# Generate dynamic content using your AI copy pipelines
	# 						initial = generate_email(lead, niche, location)
	# 						follow1 = generate_followup(lead["company"], 1)
	# 						follow2 = generate_followup(lead["company"], 2)
	# 						web_prompt = ai_generate(build_web_app_prompt(lead))
	# 						loom_script = ai_generate(build_loom_script(lead))
	# 						sms_copy = ai_generate(build_sms_copy(lead))
	# 						calendar_link = book_call(lead["company"], email)
							
	# 						profile_url = lead.get("profileUrl", f"https://x.com/{lead.get('author_id', '')}")
							
	# 						# Match the exact matrix schema of the sheet
	# 						sheet.append_row([
	# 							niche["name"],                  # Niche
	# 							location,                       # Location
	# 							lead.get("company", ""),        # Company
	# 							lead.get("website", ""),        # Website
	# 							lead.get("phone", ""),          # Phone number
	# 							email,                          # Email address
	# 							initial,                        # Initial email
	# 							follow1,                        # Follow-up 1
	# 							follow2,                        # Follow-up 2
	# 							calendar_link,                  # Calendar link
	# 							"Queued",                       # Status
	# 							today,                          # Last contacted
	# 							source_label,                   # Lead source
	# 							profile_url,                    # Profile URL
	# 							web_prompt,                     # Web prompt
	# 							loom_script,                    # Loom script
	# 							sms_copy                        # SMS copy
	# 						])
# Stripe Example usage:
# if not has_active_subscription(client["stripe_customer_id"]):
#     return


# def run_v7():
#     rows = sheet.get_all_records()

#     for i, row in enumerate(rows, start=2):
#         lead = {
#             "business": row["Name"],
#             "niche": row["Niche"],
#             "industry": row["Industry"],
#             "avatar": row["Avatar"],
#             "offer": row["Offer"],
#             "pain": row["Pain Point"],
#             "outcome": row["Desired Outcome"],
#             "size": row["Business Size"],
#             "urgency": row["Urgency"],
#             "custom": row["Custom Build"],
#         }

#         web_prompt = ai_generate(build_web_app_prompt(lead))
#         loom_script = ai_generate(build_loom_script(lead))
#         sms_copy = ai_generate(build_sms_copy(lead))
#         # price = calculate_price(lead["size"], lead["urgency"], lead["custom"])
#         # checkout = create_stripe_checkout(lead["business"], price)
#         # proposal = generate_proposal_pdf(lead, price)
#         # demo_site = generate_demo_site(lead)
#         # calendar_link = book_call(lead["business"], lead["email"])

#         # Write back to Google Sheet
#         sheet.update(f"M{i}", web_prompt)
#         sheet.update(f"N{i}", loom_script)
#         sheet.update(f"O{i}", sms_copy)
#         # sheet.update(f"P{i}", price)
#         # sheet.update(f"Q{i}", checkout)
#         # sheet.update(f"R{i}", proposal)
#         # sheet.update(f"S{i}", demo_site)
#         # sheet.update(f"T{i}", calendar_link)

#         # Push to Notion
#         push_to_notion(lead, {
#             "web_prompt": web_prompt,
#             "loom": loom_script,
#             "sms": sms_copy,
#             # "price": price,
#             # "checkout": checkout,
#             # "proposal": proposal,
#             # "demo_site": demo_site,
#             # "calendar_link": calendar_link
#         })

#         print(f"✅ Processed lead: {lead['name']}")